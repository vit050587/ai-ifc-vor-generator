"""
IFC Elements Grouping Pipeline
Input: Excel file with IFC elements
Output: Excel file with hierarchical grouping + JSON file

Поддерживает два режима:
- КР (Конструктивные решения): иерархия Часть здания → Раздел → Подраздел → Геометрия → Материал
- АР (Архитектурные решения): иерархия Часть здания → Код МССК → Наименование элемента
"""

import pandas as pd
import json
import re
import os
import time
from typing import List, Dict, Any
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from src.core.logger import setup_logger

logger = setup_logger(__name__)


# ========== Configuration ==========
GEOMETRY_GROUP_RULES = {
    'IfcWall': {
        'field': 'Ширина, мм',
        'label': 'Толщина',
        'unit': 'мм',
        'ranges': [
            {'max': 100, 'label': 'до 100 мм'},
            {'max': 150, 'label': 'до 150 мм'},
            {'max': 200, 'label': 'до 200 мм'},
            {'max': 300, 'label': 'до 300 мм'},
            {'max': float('inf'), 'label': 'более 300 мм'}
        ]
    },
    'IfcSlab': {
        'field': 'Площадь, м2',
        'label': 'Площадь',
        'unit': 'м²',
        'ranges': [
            {'max': 10, 'label': 'до 10 м²'},
            {'max': 20, 'label': 'до 20 м²'},
            {'max': float('inf'), 'label': 'более 20 м²'}
        ]
    },
    'IfcColumn': {
        'field': 'Периметр, мм',
        'label': 'Периметр',
        'unit': 'мм',
        'ranges': [
            {'max': 1200, 'label': 'до 1200 мм'},
            {'max': float('inf'), 'label': 'более 1200 мм'}
        ],
        'sub_ranges': {
            'field': 'Ширина, мм',
            'label': 'Сторона',
            'unit': 'мм',
            'ranges': [
                {'max': 300, 'label': '≤ 300 мм'},
                {'max': 500, 'label': '≤ 500 мм'},
                {'max': float('inf'), 'label': '> 500 мм'}
            ]
        }
    },
    'IfcBeam': {
        'field': 'Площадь, м2',
        'label': 'Площадь',
        'unit': 'м²',
        'ranges': [
            {'max': 10, 'label': 'до 10 м²'},
            {'max': 20, 'label': 'до 20 м²'},
            {'max': float('inf'), 'label': 'более 20 м²'}
        ]
    },
    'IfcStair': {
        'field': 'Объём, м3',
        'label': '',
        'unit': 'м³',
        'ranges': [
            {'max': float('inf'), 'label': 'Все элементы'}
        ]
    },
    'IfcStairFlight': {
        'field': 'Объём, м3',
        'label': '',
        'unit': 'м³',
        'ranges': [
            {'max': float('inf'), 'label': 'Все элементы'}
        ]
    },
    'IfcPile': {
        'field': 'Длина, мм',
        'label': 'Длина',
        'unit': 'мм',
        'ranges': [
            {'max': 8000, 'label': '6-8 м'},
            {'max': 10000, 'label': '9-10 м'},
            {'max': 12000, 'label': '11-12 м'},
            {'max': float('inf'), 'label': 'более 12 м'}
        ]
    },
    'IfcProxyElement': {
        'field': 'Объём, м3',
        'label': '',
        'unit': 'м³',
        'ranges': [
            {'max': float('inf'), 'label': 'Все элементы'}
        ]
    },
    'default': {
        'field': 'Объём, м3',
        'label': 'Объём',
        'unit': 'м³',
        'ranges': [
            {'max': 1, 'label': 'до 1 м³'},
            {'max': 5, 'label': 'до 5 м³'},
            {'max': float('inf'), 'label': 'более 5 м³'}
        ]
    }
}

NO_GEOMETRY_GROUP_SUBSECTIONS = [
    'Фундаментная плита',
    'Фундамент под инженерное оборудование',
    'Фундамент под башенный кран',
]

SECTION_STRUCTURE = {
    'Подземная': {
        'label': 'Подземная часть здания (до отм. 0,000)',
        'other_label': 'Прочие элементы подземной части',
        'sections': [
            {
                'name': 'Раздел 1. Монолитные ж/б конструкции. Фундаменты',
                'subsections': [
                    {'key': 'Фундаментная плита', 'patterns': ['фундаментная плита', 'фунд. плита', 'фундамент плита'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Свайно-ростверковый фундамент', 'patterns': ['свай', 'ростверк'], 'ifcTypes': ['IfcSlab', 'IfcBeam', 'IfcPile']},
                    {'key': 'Фундамент под инженерное оборудование', 'patterns': ['инженер', 'оборудование'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Фундамент под башенный кран', 'patterns': ['башен', 'кран'], 'ifcTypes': ['IfcSlab', 'IfcBeam']},
                    {'key': 'Устройство горизонтальной гидроизоляции', 'patterns': ['горизонт', 'гидроизол'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Устройство деформационного шва', 'patterns': ['деформац', 'шов'], 'ifcTypes': ['IfcSlab', 'IfcWall']}
                ]
            },
            {
                'name': 'Раздел 2. Монолитные ж/б конструкции. Подземная часть здания',
                'subsections': [
                    {'key': 'Подземная часть здания. Стены', 'patterns': ['стен'], 'ifcTypes': ['IfcWall']},
                    {'key': 'Подземная часть здания. Колонны', 'patterns': ['колонн'], 'ifcTypes': ['IfcColumn']},
                    {'key': 'Подземная часть здания. Плиты перекрытия', 'patterns': ['перекрыт', 'плит'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Подземная часть здания. Балки', 'patterns': ['балк', 'ригел'], 'ifcTypes': ['IfcBeam']},
                    {'key': 'Подземная часть здания. Лестницы', 'patterns': ['лестн', 'марш', 'площадк', 'плм', 'лмн'], 'ifcTypes': ['IfcStair', 'IfcSlab', 'IfcStairFlight']},
                    {'key': 'Подземная часть здания. Приямки', 'patterns': ['приям'], 'ifcTypes': ['IfcSlab', 'IfcWall']},
                    {'key': 'Подземная часть здания. Вертикальная гидроизоляция', 'patterns': ['вертикал', 'гидроизол'], 'ifcTypes': ['IfcWall']}
                ]
            }
        ]
    },
    'Цоколь': {
        'label': 'Цокольная часть здания (отм. 0,000)',
        'other_label': 'Прочие элементы цокольной части',
        'sections': [
            {
                'name': 'Раздел 1. Монолитные ж/б конструкции. Фундаменты',
                'subsections': [
                    {'key': 'Фундаментная плита', 'patterns': ['фундаментная плита', 'фунд. плита', 'фундамент плита'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Свайно-ростверковый фундамент', 'patterns': ['свай', 'ростверк'], 'ifcTypes': ['IfcSlab', 'IfcBeam', 'IfcPile']},
                    {'key': 'Фундамент под инженерное оборудование', 'patterns': ['инженер', 'оборудование'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Фундамент под башенный кран', 'patterns': ['башен', 'кран'], 'ifcTypes': ['IfcSlab', 'IfcBeam']},
                    {'key': 'Устройство горизонтальной гидроизоляции', 'patterns': ['горизонт', 'гидроизол'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Устройство вертикальной гидроизоляции', 'patterns': ['вертикал', 'гидроизол'], 'ifcTypes': ['IfcWall']},
                    {'key': 'Устройство деформационного шва', 'patterns': ['деформац', 'шов'], 'ifcTypes': ['IfcSlab', 'IfcWall']}
                ]
            },
            {
                'name': 'Раздел 2. Монолитные ж/б конструкции. Цокольная часть здания',
                'subsections': [
                    {'key': 'Цокольная часть здания. Стены', 'patterns': ['стен'], 'ifcTypes': ['IfcWall']},
                    {'key': 'Цокольная часть здания. Колонны', 'patterns': ['колонн'], 'ifcTypes': ['IfcColumn']},
                    {'key': 'Цокольная часть здания. Плиты перекрытия', 'patterns': ['перекрыт', 'плит'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Цокольная часть здания. Балки', 'patterns': ['балк', 'ригел'], 'ifcTypes': ['IfcBeam']},
                    {'key': 'Цокольная часть здания. Лестницы', 'patterns': ['лестн', 'марш', 'площадк', 'плм', 'лмн'], 'ifcTypes': ['IfcStair', 'IfcSlab', 'IfcStairFlight']},
                    {'key': 'Цокольная часть здания. Приямки', 'patterns': ['приям'], 'ifcTypes': ['IfcSlab', 'IfcWall']},
                    {'key': 'Цокольная часть здания. Вертикальная гидроизоляция', 'patterns': ['вертикал', 'гидроизол'], 'ifcTypes': ['IfcWall']}
                ]
            }
        ]
    },
    'Надземная': {
        'label': 'Надземная часть здания (выше отм. 0,000)',
        'other_label': 'Прочие элементы надземной части',
        'sections': [
            {
                'name': 'Раздел 3. Монолитные ж/б конструкции. Надземная часть здания',
                'subsections': [
                    {'key': 'Надземная часть здания. Стены', 'patterns': ['стен'], 'ifcTypes': ['IfcWall']},
                    {'key': 'Надземная часть здания. Перекрытия', 'patterns': ['перекрыт', 'плит'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Надземная часть здания. Колонны', 'patterns': ['колонн'], 'ifcTypes': ['IfcColumn']},
                    {'key': 'Надземная часть здания. Балки', 'patterns': ['балк', 'ригел'], 'ifcTypes': ['IfcBeam']},
                    {'key': 'Надземная часть здания. Парапеты', 'patterns': ['парапет'], 'ifcTypes': ['IfcWall']},
                    {'key': 'Надземная часть здания. Лестничные площадки', 'patterns': ['лестничн', 'площадк', 'плм', 'лмн'], 'ifcTypes': ['IfcSlab']},
                    {'key': 'Надземная часть здания. Лестничные марши', 'patterns': ['лестничн', 'марш'], 'ifcTypes': ['IfcStair', 'IfcStairFlight']},
                    {'key': 'Надземная часть здания. Вертикальная гидроизоляция', 'patterns': ['вертикал', 'гидроизол'], 'ifcTypes': ['IfcWall']}
                ]
            }
        ]
    }
}

IFC_TYPE_LABELS = {
    'IfcWall': 'Стены',
    'IfcWallStandardCase': 'Стены',
    'IfcSlab': 'Плиты',
    'IfcColumn': 'Колонны',
    'IfcBeam': 'Балки',
    'IfcStair': 'Лестницы',
    'IfcStairFlight': 'Лестничные марши',
    'IfcDoor': 'Двери',
    'IfcWindow': 'Окна',
    'IfcRailing': 'Ограждения',
    'IfcCovering': 'Покрытия',
    'IfcBuildingElementProxy': 'Прочие элементы',
    'IfcPile': 'Сваи',
    'IfcRamp': 'Пандусы',
}


# ========== Utility Functions ==========
def safe_parse_float(value: Any) -> float:
    if value is None or value == '' or value == '-':
        return 0.0
    try:
        cleaned = str(value).replace(' ', '').replace(',', '.')
        cleaned = re.sub(r'[^\d.\-]', '', cleaned)
        return float(cleaned) if cleaned else 0.0
    except (ValueError, TypeError):
        return 0.0


def round_value(value: float, column_name: str) -> float:
    col_lower = column_name.lower()
    if 'м3' in col_lower or 'литр' in col_lower:
        return round(value, 3)
    elif 'м2' in col_lower:
        return round(value, 2)
    elif 'мм' in col_lower:
        return round(value)
    else:
        return round(value, 2)


def clean_element_name(name: str) -> str:
    if not name or not isinstance(name, str):
        return 'Без названия'
    cleaned = re.sub(r':\s*\d+$', '', name).strip()
    return cleaned or name.strip()


# ========== Classification Functions ==========
def get_ifc_type(type_ru: str, name: str) -> str:
    name_lower = name.lower() if name else ''
    type_lower = type_ru.lower() if type_ru else ''
    
    if 'ifcproxy' in type_lower or 'proxy' in type_lower:
        return 'IfcProxyElement'
    
    if any(w in name_lower for w in ['отверсти', 'проём', 'проем', 'окно', 'двер']):
        return 'IfcOpening'
    
    if 'труб' in name_lower:
        return 'IfcProxyElement'
    
    if 'ifcpile' in type_lower:
        return 'IfcPile'
    if 'ifcstairflight' in type_lower:
        return 'IfcStairFlight'
    
    if 'сва' in name_lower:
        return 'IfcPile'
    
    if any(w in name_lower for w in ['плм', 'лмн', 'площадк']):
        return 'IfcSlab'
    
    if any(w in name_lower for w in ['стен', 'парапет']):
        return 'IfcWall'
    if any(w in name_lower for w in ['колонн', 'пилон']):
        return 'IfcColumn'
    if any(w in name_lower for w in ['балк', 'ригел', 'перемычк']):
        return 'IfcBeam'
    if any(w in name_lower for w in ['лестн', 'марш']):
        return 'IfcStair'
    if 'площадк' in name_lower:
        return 'IfcSlab'
    if any(w in name_lower for w in ['перекрыт', 'плит', 'приям', 'фундамент']):
        return 'IfcSlab'
    
    if 'ifcwall' in type_lower:
        return 'IfcWall'
    if 'ifcslab' in type_lower:
        return 'IfcSlab'
    if 'ifccolumn' in type_lower:
        return 'IfcColumn'
    if 'ifcbeam' in type_lower:
        return 'IfcBeam'
    if 'ifcstair' in type_lower:
        return 'IfcStair'
    
    return 'default'


def is_hydro_vertical(type_ru: str, name: str, ifc_type: str) -> bool:
    name_lower = name.lower() if name else ''
    has_hydro = 'гидроизол' in name_lower
    has_vertical = 'вертикал' in name_lower
    has_wall = 'стен' in name_lower or ifc_type == 'IfcWall'
    has_fundament = any(w in name_lower for w in ['фундамент', 'фунд. плита', 'фундаментная'])
    return has_hydro and has_wall and not has_fundament


def is_hydro_horizontal(type_ru: str, name: str, ifc_type: str) -> bool:
    name_lower = name.lower() if name else ''
    has_hydro = 'гидроизол' in name_lower
    has_fundament_or_slab = any(w in name_lower for w in ['фундамент', 'фунд. плита', 'фундаментная', 'плит', 'перекрыт']) or ifc_type == 'IfcSlab'
    if is_hydro_vertical(type_ru, name, ifc_type):
        return False
    return has_hydro and has_fundament_or_slab


def determine_subsection(type_ru: str, name: str, part: str) -> str:
    name_lower = name.lower() if name else ''
    type_lower = type_ru.lower() if type_ru else ''
    
    temp_ifc_type = get_ifc_type(type_ru, name)
    
    if 'ifcproxy' in type_lower or 'proxy' in type_lower:
        return '__OTHER__'
    
    if temp_ifc_type == 'IfcOpening':
        return '__OTHER__'
    
    if 'труб' in name_lower:
        return '__OTHER__'
    
    if is_hydro_vertical(type_ru, name, temp_ifc_type):
        if part == 'Подземная':
            return 'Подземная часть здания. Вертикальная гидроизоляция'
        elif part == 'Цоколь':
            return 'Цокольная часть здания. Вертикальная гидроизоляция'
        else:
            return 'Надземная часть здания. Вертикальная гидроизоляция'
    if is_hydro_horizontal(type_ru, name, temp_ifc_type):
        return 'Устройство горизонтальной гидроизоляции'
    
    if 'сва' in name_lower or 'ifcpile' in type_lower:
        return 'Свайно-ростверковый фундамент'
    
    if 'приям' in name_lower:
        mapping = {'Подземная': 'Подземная часть здания. Приямки',
                   'Цоколь': 'Цокольная часть здания. Приямки'}
        return mapping.get(part, 'Надземная часть здания. Приямки')
    
    if 'парапет' in name_lower:
        return 'Надземная часть здания. Парапеты'
    
    if temp_ifc_type == 'IfcStairFlight':
        if part in ('Подземная', 'Цоколь'):
            return f'{part} часть здания. Лестницы'
        return 'Надземная часть здания. Лестничные марши'
    
    if 'плм' in name_lower:
        if part in ('Подземная', 'Цоколь'):
            return f'{part} часть здания. Лестницы'
        return 'Надземная часть здания. Лестничные площадки'
    
    if 'лмн' in name_lower:
        if part in ('Подземная', 'Цоколь'):
            return f'{part} часть здания. Лестницы'
        return 'Надземная часть здания. Лестничные площадки'
    
    if 'лестничн' in name_lower or 'лестниц' in name_lower:
        if 'площадк' in name_lower:
            mapping = {'Подземная': 'Подземная часть здания. Лестницы',
                       'Цоколь': 'Цокольная часть здания. Лестницы'}
            return mapping.get(part, 'Надземная часть здания. Лестничные площадки')
        if 'марш' in name_lower:
            mapping = {'Подземная': 'Подземная часть здания. Лестницы',
                       'Цоколь': 'Цокольная часть здания. Лестницы'}
            return mapping.get(part, 'Надземная часть здания. Лестничные марши')
        mapping = {'Подземная': 'Подземная часть здания. Лестницы',
                   'Цоколь': 'Цокольная часть здания. Лестницы'}
        return mapping.get(part, 'Надземная часть здания. Лестничные марши')
    
    part_structure = SECTION_STRUCTURE.get(part)
    if part_structure:
        for section in part_structure['sections']:
            for subsection in section['subsections']:
                if any(pattern in name_lower for pattern in subsection['patterns']):
                    return subsection['key']
    
    type_mapping = {
        'wall': 'Стены',
        'slab': 'Плиты перекрытия',
        'column': 'Колонны',
        'beam': 'Балки',
        'stair': 'Лестницы'
    }
    for key, value in type_mapping.items():
        if key in type_lower:
            return f'{part} часть здания. {value}'
    
    return '__OTHER__'


# ========== Core Grouping Logic (КР) ==========
class ElementData:
    def __init__(self, index: int, row: Dict[str, Any], headers: List[str]):
        self.index = index
        self.row = row
        type_ru_idx = headers.index('Тип элемента') if 'Тип элемента' in headers else -1
        name_idx = headers.index('Имя') if 'Имя' in headers else -1
        self.type_ru = row.get(headers[type_ru_idx], 'Неизвестно') if type_ru_idx >= 0 else 'Неизвестно'
        self.name = row.get(headers[name_idx], '') if name_idx >= 0 else ''
        self.part = row.get('Часть здания', 'Надземная')
        self.subsection = determine_subsection(self.type_ru, self.name, self.part)
        self.ifc_type = get_ifc_type(self.type_ru, self.name)


def _create_group(name: str, level: int, group_df, rows: List[Dict[str, Any]],
                  volume_col, sum_columns: List[str], children=None) -> Dict[str, Any]:
    """Создаёт узел дерева групп: агрегирует объём/площади по DataFrame строк.

    Вынесен на уровень модуля, чтобы переиспользовать в разных вариантах
    группировки (по разделам/подразделам и по кодам МССК).
    """
    indices = sorted(group_df.index.tolist())

    if volume_col and volume_col in group_df.columns:
        vol_series = pd.to_numeric(group_df[volume_col], errors='coerce').fillna(0)
        # float() — иначе np.int64/np.float64 попадут в JSON как строки
        volume = float(round(vol_series.sum(), 2))
    else:
        volume = 0.0

    areas = {}
    for col in sum_columns:
        if 'площадь' in col.lower() and col in group_df.columns:
            area_series = pd.to_numeric(group_df[col], errors='coerce').fillna(0)
            total = area_series.sum()
            if total > 0:
                areas[col] = float(round(total, 2))

    first_elem = rows[indices[0]] if indices else {}

    return {
        'name': name, 'level': level, 'indices': indices,
        'total_volume': volume, 'total_areas': areas,
        'first_element': dict(first_elem), 'count': len(indices),
        'children': children or []
    }


def group_elements(rows: List[Dict[str, Any]], headers: List[str], use_new_grouping: bool = False) -> List[Dict[str, Any]]:
    """Группировка элементов по иерархии (для КР). Оптимизированная версия с pandas.
    
    use_new_grouping=True — упрощённая геометрическая группировка из group_excel_new.py.
    """
    
    volume_col = None
    for h in headers:
        h_lower = h.lower()
        if 'netvolume' in h_lower or ('объём' in h_lower and 'м3' in h_lower):
            volume_col = h
            break
    
    sum_columns = [h for h in headers if any(kw in h.lower() for kw in ['объём', 'объем', 'площадь', 'стоимость'])]
    
    df = pd.DataFrame(rows)
    
    if volume_col and volume_col in df.columns:
        df[volume_col] = pd.to_numeric(df[volume_col], errors='coerce').fillna(0)
    
    def get_part(floor_type):
        if not floor_type or pd.isna(floor_type):
            return 'Надземная'
        fl = str(floor_type).lower().strip()
        if any(w in fl for w in ['подзем', 'подвал', 'basement', '-1']):
            return 'Подземная'
        if any(w in fl for w in ['цокол', 'ground', 'нулев']):
            return 'Цоколь'
        if any(w in fl for w in ['надзем', 'этаж', 'кровл', 'техническ', 'мансард']):
            return 'Надземная'
        return 'Надземная'
    
    def get_subsection(row):
        type_ru = str(row.get('Тип элемента', ''))
        name = str(row.get('Имя', ''))
        part = row['_part']
        return determine_subsection(type_ru, name, part)
    
    def get_ifc_type_row(row):
        return get_ifc_type(str(row.get('Тип элемента', '')), str(row.get('Имя', '')))
    
    df['_part'] = df['Тип_этажа'].apply(get_part) if 'Тип_этажа' in df.columns else 'Надземная'
    df['_subsection'] = df.apply(get_subsection, axis=1)
    df['_ifc_type'] = df.apply(get_ifc_type_row, axis=1)
    
    def create_group(name, level, group_df, children=None):
        return _create_group(name, level, group_df, rows, volume_col, sum_columns, children)
    
    def get_material_group(row):
        material = str(row.get('Материал', ''))
        if not material or material == '-' or material == '' or pd.isna(row.get('Материал')):
            return 'Материал: не указан'
        return f'Материал: {material}'
    
    def get_concrete_group(row):
        concrete = str(row.get('ExpCheck_MaterialConcrete_MGE_ConcreteGrade', ''))
        water = str(row.get('ExpCheck_MaterialConcrete_MGE_WaterResist', ''))
        freeze = str(row.get('ExpCheck_MaterialConcrete_MGE_FreezeDurability', ''))
        
        parts = []
        if concrete and concrete != '-' and concrete != '' and concrete != 'nan':
            parts.append(concrete)
        if water and water != '-' and water != '' and water != 'nan':
            parts.append(f"W{water}" if not water.startswith('W') else water)
        if freeze and freeze != '-' and freeze != '' and freeze != 'nan':
            parts.append(f"F{freeze}" if not freeze.startswith('F') else freeze)
        
        return f"Бетон: {', '.join(parts)}" if parts else "Бетон: без характеристик"
    
    def get_name_key(row):
        name = str(row.get('Имя', '')).strip()
        if not name:
            return 'Без названия'
        last_colon = name.rfind(':')
        if last_colon != -1:
            after = name[last_colon + 1:].strip()
            if after.isdigit():
                name = name[:last_colon].strip()
        return name if name else 'Без названия'
    
    def apply_material_concrete(group_df, parent_group, mat_level, concrete_level):
        if len(group_df) == 0:
            return
        
        group_df = group_df.copy()
        group_df['_mat'] = group_df.apply(get_material_group, axis=1)
        mat_groups = group_df.groupby('_mat')
        
        if len(mat_groups) == 1:
            group_df['_conc'] = group_df.apply(get_concrete_group, axis=1)
            conc_groups = group_df.groupby('_conc')
            
            if len(conc_groups) > 1:
                for conc_name, conc_df in conc_groups:
                    g = create_group(conc_name, concrete_level, conc_df)
                    if g:
                        parent_group['children'].append(g)
        else:
            for mat_name, mat_df in mat_groups:
                mat_group = create_group(mat_name, mat_level, mat_df)
                if mat_group:
                    mat_df = mat_df.copy()
                    mat_df['_conc'] = mat_df.apply(get_concrete_group, axis=1)
                    conc_groups = mat_df.groupby('_conc')
                    
                    if len(conc_groups) > 1:
                        for conc_name, conc_df in conc_groups:
                            g = create_group(conc_name, concrete_level, conc_df)
                            if g:
                                mat_group['children'].append(g)
                    
                    parent_group['children'].append(mat_group)
    
    result = []
    part_order = ['Подземная', 'Цоколь', 'Надземная']
    
    for part in part_order:
        part_df = df[df['_part'] == part]
        if len(part_df) == 0:
            continue
        
        part_structure = SECTION_STRUCTURE.get(part)
        if not part_structure:
            continue
        
        part_group = create_group(part_structure['label'], 1, part_df)
        
        other_mask = part_df['_subsection'].isin(['__OTHER__']) | part_df['_ifc_type'].isin(['IfcProxyElement', 'IfcOpening'])
        regular_df = part_df[~other_mask]
        other_df = part_df[other_mask]
        
        for section in part_structure['sections']:
            section_keys = [sub['key'] for sub in section['subsections']]
            section_df = regular_df[regular_df['_subsection'].isin(section_keys)]
            
            if len(section_df) == 0:
                continue
            
            section_group = create_group(section['name'], 2, section_df)
            
            for subsection in section['subsections']:
                sub_df = section_df[section_df['_subsection'] == subsection['key']]
                if len(sub_df) == 0:
                    continue
                
                sub_group = create_group(subsection['key'], 3, sub_df)
                
                skip_geometry = subsection['key'] in NO_GEOMETRY_GROUP_SUBSECTIONS
                
                if skip_geometry:
                    apply_material_concrete(sub_df, sub_group, 4, 5)
                else:
                    elems = []
                    for idx in sub_df.index:
                        row = rows[idx]
                        e = ElementData(idx, row, headers)
                        e.part = part
                        e.subsection = subsection['key']
                        e.ifc_type = get_ifc_type(str(row.get('Тип элемента', '')), str(row.get('Имя', '')))
                        elems.append(e)
                    
                    _add_geometry_groups(sub_group, elems, headers, volume_col, use_new_grouping=use_new_grouping)
                
                if sub_group['children']:
                    section_group['children'].append(sub_group)
            
            if section_group['children']:
                part_group['children'].append(section_group)
        
        if len(other_df) > 0:
            other_group = create_group(part_structure.get('other_label', 'Прочие элементы'), 2, other_df)
            
            other_df = other_df.copy()
            other_df['_name_key'] = other_df.apply(get_name_key, axis=1)
            for name_key, name_df in other_df.groupby('_name_key'):
                name_group = create_group(name_key, 3, name_df)
                apply_material_concrete(name_df, name_group, 4, 5)
                if name_group['children']:
                    other_group['children'].append(name_group)
            
            if other_group['children']:
                part_group['children'].append(other_group)
        
        result.append(part_group)
    
    return result


def group_elements_new(rows: List[Dict[str, Any]], headers: List[str]) -> List[Dict[str, Any]]:
    """Группировка элементов (для КР) с упрощённой геометрической логикой из group_excel_new.py.

    Отличия от group_elements:
    - фиксированные уровни: ifc-группа = 4, геометрические/подгруппы = 5;
    - подгруппы sub_ranges добавляются напрямую в родителя (без промежуточной группы);
    - float()-приведение значений объёмов/площадей для корректной JSON-сериализации.
    """
    return group_elements(rows, headers, use_new_grouping=True)


def group_elements_mssk(rows: List[Dict[str, Any]], headers: List[str]) -> List[Dict[str, Any]]:
    """Группировка элементов (для КР) по иерархии:

        Часть здания (L1)
          → Код МССК (L2) — название из справочника elements_mssk_nested.json,
                             неизвестные/пустые коды объединяются в «Прочее»
            → IFC-тип (L3, только если в МССК-группе несколько типов)
              → Геометрия (L4) — по GEOMETRY_GROUP_RULES:
                    стены — Ширина, мм (толщина),
                    плиты/балки — Площадь, м2,
                    колонны — Периметр, мм + подгруппа «Сторона» по ширине,
                    сваи — Длина, мм и т.д.
                → Материал (L5) — «Материал: …» (только при нескольких материалах)
                  → Бетон (L6) — «Бетон: В…, W…, F…»
                       (по колонкам ExpCheck_MaterialConcrete_*)

    Геометрическая группировка и материал/бетон переиспользуют _add_geometry_groups
    со сдвигом уровней level_offset=-1.
    """
    # Импорт внутри функции, чтобы избежать циклической зависимости
    from src.services.mssk_lookup import build_mssk_lookup, OTHER_LABEL
    lookup, _ = build_mssk_lookup()

    volume_col = None
    for h in headers:
        h_lower = h.lower()
        if 'netvolume' in h_lower or ('объём' in h_lower and 'м3' in h_lower):
            volume_col = h
            break

    sum_columns = [h for h in headers if any(kw in h.lower() for kw in ['объём', 'объем', 'площадь', 'стоимость'])]

    df = pd.DataFrame(rows)
    if volume_col and volume_col in df.columns:
        df[volume_col] = pd.to_numeric(df[volume_col], errors='coerce').fillna(0)

    def get_part(floor_type):
        if not floor_type or pd.isna(floor_type):
            return 'Надземная'
        fl = str(floor_type).lower().strip()
        if any(w in fl for w in ['подзем', 'подвал', 'basement', '-1']):
            return 'Подземная'
        if any(w in fl for w in ['цокол', 'ground', 'нулев']):
            return 'Цоколь'
        if any(w in fl for w in ['надзем', 'этаж', 'кровл', 'техническ', 'мансард']):
            return 'Надземная'
        return 'Надземная'

    def get_ifc_type_row(row):
        return get_ifc_type(str(row.get('Тип элемента', '')), str(row.get('Имя', '')))

    df['_part'] = df['Тип_этажа'].apply(get_part) if 'Тип_этажа' in df.columns else 'Надземная'
    df['_ifc_type'] = df.apply(get_ifc_type_row, axis=1)

    has_mssk_col = 'Код мсск' in headers

    result = []
    part_order = ['Подземная', 'Цоколь', 'Надземная']
    part_labels = {
        'Подземная': 'Подземная часть здания (до отм. 0,000)',
        'Цоколь': 'Цокольная часть здания (отм. 0,000)',
        'Надземная': 'Надземная часть здания (выше отм. 0,000)',
    }

    for part in part_order:
        part_df = df[df['_part'] == part]
        if len(part_df) == 0:
            continue

        part_group = _create_group(part_labels[part], 1, part_df, rows, volume_col, sum_columns)

        # --- Группировка по коду МССК ---
        mssk_groups = defaultdict(list)
        mssk_meta = {}  # key → (название_группы, порядок_сортировки)
        for idx in part_df.index:
            raw_code = rows[idx].get('Код мсск', '') if has_mssk_col else ''
            code = str(raw_code).strip() if raw_code is not None else ''
            if code and code != '-':
                info = lookup.get(code)
                if info:
                    key = f'{code}__{info["name"]}'
                    meta = (info['name'], info['order'])
                else:
                    key = '__OTHER__'
                    meta = (OTHER_LABEL, float('inf'))
            else:
                key = '__OTHER__'
                meta = (OTHER_LABEL, float('inf'))
            mssk_groups[key].append(idx)
            mssk_meta[key] = meta

        # Сортировка: сначала известные коды (по порядку в справочнике),
        # затем «Прочее»
        sorted_keys = sorted(mssk_groups.keys(),
                             key=lambda k: (mssk_meta[k][1], mssk_meta[k][0]))

        for key in sorted_keys:
            indices = sorted(mssk_groups[key])
            mssk_df = df.loc[indices]
            name = mssk_meta[key][0]

            mssk_group = _create_group(name, 2, mssk_df, rows, volume_col, sum_columns)

            # --- Геометрия + Материал/Бетон ---
            elems = []
            for idx in indices:
                row = rows[idx]
                e = ElementData(idx, row, headers)
                e.part = part
                e.ifc_type = get_ifc_type(str(row.get('Тип элемента', '')), str(row.get('Имя', '')))
                elems.append(e)

            # level_offset=-1: ifc=3, геометрия=4, материал=5, бетон=6;
            # для колонн с подгруппами: «Периметр»=4 → «Сторона»=5 → материал=6 → бетон=7
            # nest_sub_ranges=True: для колонн — «Периметр» → «Сторона» (вложенно)
            _add_geometry_groups(mssk_group, elems, headers, volume_col,
                                 use_new_grouping=True, level_offset=-1,
                                 nest_sub_ranges=True)

            if mssk_group['children']:
                part_group['children'].append(mssk_group)

        result.append(part_group)

    return result


def _add_geometry_groups(parent_group, elems, headers, volume_col, use_new_grouping=False,
                         level_offset=0, nest_sub_ranges=False):
    """Добавляет геометрическую группировку к родительской группе с поддержкой sub_ranges.

    use_new_grouping=True — упрощённая логика из group_excel_new.py:
    фиксированные уровни (ifc=4, геометрия/подгруппы=5), подгруппы sub_ranges
    добавляются напрямую в родителя.
    use_new_grouping=False (по умолчанию) — старая логика: динамический base_level,
    промежуточная основная группа с вложенными подгруппами.

    level_offset — сдвиг уровней для иерархий другой глубины (например,
    группировка по кодам МССК: ifc=3, геометрия=4, материал=5, бетон=6).

    nest_sub_ranges=True (только вместе с use_new_grouping=True) — сохраняет
    промежуточную основную геометрическую группу (например «Периметр: до 1200 мм»)
    и вкладывает подгруппы sub_ranges («Сторона: ≤ 300 мм») внутрь неё.
    """
    mat_level = 6 + level_offset
    conc_level = 7 + level_offset
    # Уровни геометрии для use_new_grouping=True с nest_sub_ranges:
    # основная группа на уровне геометрии, подгруппы — на уровень глубже
    geo_level = 5 + level_offset
    sub_level = 6 + level_offset

    def get_volume(elem):
        if volume_col is None:
            return 0.0
        return safe_parse_float(elem.row.get(volume_col, 0))
    
    def get_geometry_value(rule, elem):
        if not rule or not rule.get('field'):
            return get_volume(elem)
        field = rule['field']
        if field not in headers:
            return get_volume(elem)
        return safe_parse_float(elem.row.get(field, 0))
    
    def create_geo_group(name, level, elems_list):
        if not elems_list:
            return None
        volume = sum(get_volume(e) for e in elems_list)
        indices = sorted([e.index for e in elems_list])
        return {
            'name': name, 'level': level, 'indices': indices,
            # НОВАЯ ВЕРСИЯ: float() — чтобы значения корректно сериализовались в JSON
            'total_volume': float(round(volume, 2)), 'total_areas': {},
            'first_element': dict(elems_list[0].row) if elems_list else {},
            'count': len(elems_list), 'children': []
        }
    
    def apply_mat_conc(elems_list, parent, mat_level=mat_level, conc_level=conc_level):
        """Группировка по материалу/бетону для геометрических групп.

        Сначала «Материал: …», затем внутри — «Бетон: В…, W…, F…»
        (по колонкам ExpCheck_MaterialConcrete_*).
        Уровни можно переопределить для вложенных подгрупп (sub_ranges).
        """
        mat_groups = defaultdict(list)
        for e in elems_list:
            material = str(e.row.get('Материал', ''))
            mat_key = f'Материал: {material}' if material and material != '-' else 'Материал: не указан'
            mat_groups[mat_key].append(e)
        
        if len(mat_groups) == 1:
            conc_groups = defaultdict(list)
            for e in elems_list:
                conc_key = _get_concrete_key(e)
                conc_groups[conc_key].append(e)
            if len(conc_groups) > 1:
                for conc_key, conc_elems in conc_groups.items():
                    g = create_geo_group(conc_key, conc_level, conc_elems)
                    if g:
                        parent['children'].append(g)
        else:
            for mat_key, mat_elems in mat_groups.items():
                mat_group = create_geo_group(mat_key, mat_level, mat_elems)
                if not mat_group:
                    continue
                conc_groups = defaultdict(list)
                for e in mat_elems:
                    conc_key = _get_concrete_key(e)
                    conc_groups[conc_key].append(e)
                if len(conc_groups) > 1:
                    for conc_key, conc_elems in conc_groups.items():
                        g = create_geo_group(conc_key, conc_level, conc_elems)
                        if g:
                            mat_group['children'].append(g)
                parent['children'].append(mat_group)
    
    def _get_concrete_key(elem):
        concrete = str(elem.row.get('ExpCheck_MaterialConcrete_MGE_ConcreteGrade', ''))
        water = str(elem.row.get('ExpCheck_MaterialConcrete_MGE_WaterResist', ''))
        freeze = str(elem.row.get('ExpCheck_MaterialConcrete_MGE_FreezeDurability', ''))
        parts = []
        if concrete and concrete != '-' and concrete != '' and concrete != 'nan':
            parts.append(concrete)
        if water and water != '-' and water != '' and water != 'nan':
            parts.append(f"W{water}" if not water.startswith('W') else water)
        if freeze and freeze != '-' and freeze != '' and freeze != 'nan':
            parts.append(f"F{freeze}" if not freeze.startswith('F') else freeze)
        return f"Бетон: {', '.join(parts)}" if parts else "Бетон: без характеристик"
    
    ifc_groups = {}
    for e in elems:
        if e.ifc_type not in ifc_groups:
            ifc_groups[e.ifc_type] = []
        ifc_groups[e.ifc_type].append(e)
    
    need_ifc_group = len(ifc_groups) > 1
    
    for ifc_type, ifc_elements in ifc_groups.items():
        rule = GEOMETRY_GROUP_RULES.get(ifc_type, GEOMETRY_GROUP_RULES['default'])
        current_parent = parent_group
        base_level = 4
        
        if need_ifc_group:
            ifc_labels = {
                'IfcWall': 'Стены', 'IfcSlab': 'Плиты',
                'IfcColumn': 'Колонны', 'IfcBeam': 'Балки',
                'IfcStair': 'Лестницы', 'IfcStairFlight': 'Лестничные марши',
                'IfcPile': 'Сваи', 'default': 'Прочее'
            }
            ifc_group = create_geo_group(ifc_labels.get(ifc_type, ifc_type), base_level + level_offset, ifc_elements)
            if ifc_group:
                parent_group['children'].append(ifc_group)
                current_parent = ifc_group
                # НОВАЯ ВЕРСИЯ: фиксированные уровни (4/5), base_level не увеличиваем
                if not use_new_grouping:
                    base_level += 1
        
        # Шаг 1: Основная группировка по ranges
        geo_groups = {}
        for rg in rule['ranges']:
            geo_groups[rg['label']] = []
        
        for e in ifc_elements:
            value = get_geometry_value(rule, e)
            assigned = False
            for rg in rule['ranges']:
                if value <= rg['max']:
                    geo_groups[rg['label']].append(e)
                    assigned = True
                    break
            if not assigned:
                geo_groups[rule['ranges'][-1]['label']].append(e)
        
        # Шаг 2: Создание групп с учётом sub_ranges
        for geo_label, geo_elements in geo_groups.items():
            if not geo_elements:
                continue
            
            rule_label = str(rule.get('label', '')).strip()
            geo_name = f'{rule_label}: {geo_label}' if rule_label else geo_label
            
            # Если есть sub_ranges и элементов больше 1 (старая логика)
            if rule.get('sub_ranges') and len(geo_elements) > 1:
                sub_rule = rule['sub_ranges']
                sub_geo_groups = defaultdict(list)
                
                # Распределяем по sub_ranges
                for e in geo_elements:
                    val = safe_parse_float(e.row.get(sub_rule['field'], 0))
                    assigned = False
                    for rg in sub_rule['ranges']:
                        if val <= rg['max']:
                            sub_geo_groups[rg['label']].append(e)
                            assigned = True
                            break
                    if not assigned:
                        sub_geo_groups[sub_rule['ranges'][-1]['label']].append(e)
                
                if use_new_grouping:
                    if nest_sub_ranges:
                        # Промежуточная основная группа (например «Периметр: до 1200 мм»)
                        # с вложенными подгруппами sub_ranges («Сторона: ≤ 300 мм»),
                        # фиксированные уровни: основная = geo_level, подгруппа = sub_level
                        main_geo_group = create_geo_group(geo_name, geo_level, geo_elements)
                        if not main_geo_group:
                            continue

                        has_subgroups = False
                        for sub_label, sub_elems in sub_geo_groups.items():
                            if not sub_elems:
                                continue
                            sub_name = f'{sub_rule["label"]}: {sub_label}'
                            sub_geo_group = create_geo_group(sub_name, sub_level, sub_elems)
                            if sub_geo_group:
                                apply_mat_conc(sub_elems, sub_geo_group,
                                               mat_level=sub_level + 1, conc_level=sub_level + 2)
                                main_geo_group['children'].append(sub_geo_group)
                                has_subgroups = True

                        if has_subgroups:
                            current_parent['children'].append(main_geo_group)
                        else:
                            apply_mat_conc(geo_elements, main_geo_group)
                            current_parent['children'].append(main_geo_group)
                    else:
                        # НОВАЯ ВЕРСИЯ: подгруппы добавляются напрямую в родителя
                        # (без промежуточной основной группы), уровень фиксированный
                        for sub_label, sub_elems in sub_geo_groups.items():
                            if not sub_elems:
                                continue
                            sub_geo_group = create_geo_group(f'{sub_rule["label"]}: {sub_label}', 5 + level_offset, sub_elems)
                            if sub_geo_group:
                                apply_mat_conc(sub_elems, sub_geo_group)
                                current_parent['children'].append(sub_geo_group)
                else:
                    # СТАРАЯ ВЕРСИЯ: промежуточная основная группа с вложенными подгруппами
                    main_geo_group = create_geo_group(geo_name, base_level, geo_elements)
                    if not main_geo_group:
                        continue
                    
                    # Создаём подгруппы
                    has_subgroups = False
                    for sub_label, sub_elems in sub_geo_groups.items():
                        if not sub_elems:
                            continue
                        
                        sub_name = f'{sub_rule["label"]}: {sub_label}'
                        sub_geo_group = create_geo_group(sub_name, base_level + 1, sub_elems)
                        if sub_geo_group:
                            apply_mat_conc(sub_elems, sub_geo_group)
                            main_geo_group['children'].append(sub_geo_group)
                            has_subgroups = True
                    
                    # Добавляем основную группу, только если есть подгруппы
                    if has_subgroups:
                        current_parent['children'].append(main_geo_group)
                    else:
                        # Если подгруппы не создались — добавляем как обычную группу
                        apply_mat_conc(geo_elements, main_geo_group)
                        current_parent['children'].append(main_geo_group)
            else:
                # Обычная группа без sub_ranges
                # НОВАЯ ВЕРСИЯ: фиксированный уровень, старая: base_level
                level = (5 + level_offset) if use_new_grouping else base_level
                geo_group = create_geo_group(geo_name, level, geo_elements)
                if geo_group:
                    apply_mat_conc(geo_elements, geo_group)
                    current_parent['children'].append(geo_group)

# ========== Core Grouping Logic (АР) ==========
def group_elements_ar(rows: List[Dict[str, Any]], headers: List[str]) -> List[Dict[str, Any]]:
    """
    Группировка элементов для Архитектурных Решений.
    Иерархия: Часть здания → Код МССК → Материал → Наименование элемента

    После группировки по коду МССК элементы дополнительно группируются
    по главному материалу (уровень L3), а затем — по наименованию (L4).

    Материал берётся из колонки вида «Свойство::IfcMaterialLayer::Name»
    (значения вида «Минераловатная плита (СТ 10 14 20 14)»). Имя группы
    определяется по коду МССК материала из data/materials_mssk_nested.json:
      * код не найден / поле пустое  → «Прочее»
      * несколько материалов в поле  → «Многослойные»

    Если колонка материалов отсутствует во входных данных — группировка
    выполняется как раньше: Код МССК → Наименование (без уровня материала).
    """
    from src.services.mssk_lookup import build_mssk_lookup, OTHER_LABEL as ELEMENTS_OTHER_LABEL
    from src.services.materials_lookup import (
        build_materials_lookup,
        resolve_material_group,
    )
    lookup, _ = build_mssk_lookup()
    materials_lookup, _ = build_materials_lookup()

    df = pd.DataFrame(rows)

    volume_col_name = None
    for h in headers:
        h_lower = h.lower()
        if 'netvolume' in h_lower or ('объём' in h_lower and 'м3' in h_lower):
            volume_col_name = h
            break

    sum_columns = [h for h in headers if any(kw in h.lower() for kw in ['объём', 'объем', 'площадь', 'стоимость'])]

    if volume_col_name and volume_col_name in df.columns:
        df[volume_col_name] = pd.to_numeric(df[volume_col_name], errors='coerce').fillna(0)

    # Колонка с материалами слоёв (например «Свойство::IfcMaterialLayer::Name»)
    material_col = next(
        (h for h in headers if 'IfcMaterialLayer::Name' in str(h)),
        None,
    )

    def get_part(floor_type):
        if not floor_type or pd.isna(floor_type):
            return 'Надземная'
        fl = str(floor_type).lower().strip()
        if any(w in fl for w in ['подзем', 'подвал', 'basement', '-1']):
            return 'Подземная'
        if any(w in fl for w in ['цокол', 'ground', 'нулев']):
            return 'Цоколь'
        return 'Надземная'

    if 'Тип_этажа' in df.columns:
        df['_part'] = df['Тип_этажа'].apply(get_part)
    else:
        df['_part'] = 'Надземная'

    has_mssk_col = 'Код мсск' in headers

    result = []
    part_order = ['Подземная', 'Цоколь', 'Надземная']
    part_labels = {
        'Подземная': 'Подземная часть здания (до отм. 0,000)',
        'Цоколь': 'Цокольная часть здания (отм. 0,000)',
        'Надземная': 'Надземная часть здания (выше отм. 0,000)',
    }

    def add_name_groups(parent_group, idx_list, base_level):
        """Добавляет в parent_group дочерние группы по наименованию элемента."""
        name_groups = defaultdict(list)
        for idx in idx_list:
            name_key = clean_element_name(str(rows[idx].get('Имя', '')).strip())
            name_groups[name_key].append(idx)

        for name_key in sorted(name_groups.keys()):
            name_indices = sorted(name_groups[name_key])
            name_df = df.loc[name_indices]
            name_group = _create_group(name_key, base_level, name_df, rows, volume_col_name, sum_columns)
            if name_group:
                parent_group['children'].append(name_group)

    for part in part_order:
        part_df = df[df['_part'] == part]
        if len(part_df) == 0:
            continue

        part_group = _create_group(part_labels[part], 1, part_df, rows, volume_col_name, sum_columns)

        # --- Группировка по коду МССК (L2) ---
        mssk_groups = defaultdict(list)
        mssk_meta = {}  # key → (название_группы, порядок_сортировки)
        for idx in part_df.index:
            raw_code = rows[idx].get('Код мсск', '') if has_mssk_col else ''
            code = str(raw_code).strip() if raw_code is not None else ''
            if code and code != '-':
                info = lookup.get(code)
                if info:
                    key = f'{code}__{info["name"]}'
                    meta = (info['name'], info['order'])
                else:
                    key = '__OTHER__'
                    meta = (ELEMENTS_OTHER_LABEL, float('inf'))
            else:
                key = '__OTHER__'
                meta = (ELEMENTS_OTHER_LABEL, float('inf'))
            mssk_groups[key].append(idx)
            mssk_meta[key] = meta

        # Сортировка: сначала известные коды (по порядку в справочнике),
        # затем «Прочее»
        sorted_keys = sorted(mssk_groups.keys(),
                             key=lambda k: (mssk_meta[k][1], mssk_meta[k][0]))

        for key in sorted_keys:
            indices = sorted(mssk_groups[key])
            mssk_df = df.loc[indices]
            mssk_name = mssk_meta[key][0]

            mssk_group = _create_group(mssk_name, 2, mssk_df, rows, volume_col_name, sum_columns)

            # --- Группировка по главному материалу (L3) ---
            if material_col:
                mat_groups = defaultdict(list)
                mat_meta = {}  # имя группы → порядок сортировки
                for idx in indices:
                    mat_val = rows[idx].get(material_col, '')
                    mat_name, mat_order = resolve_material_group(mat_val, materials_lookup)
                    mat_groups[mat_name].append(idx)
                    mat_meta[mat_name] = mat_order

                sorted_mat_names = sorted(mat_groups.keys(),
                                          key=lambda n: (mat_meta[n], n))

                for mat_name in sorted_mat_names:
                    mat_indices = sorted(mat_groups[mat_name])
                    mat_df = df.loc[mat_indices]
                    mat_group = _create_group(mat_name, 3, mat_df, rows, volume_col_name, sum_columns)

                    # --- Группировка по наименованию элемента (L4) ---
                    add_name_groups(mat_group, mat_indices, 4)

                    if mat_group['children']:
                        mssk_group['children'].append(mat_group)
            else:
                # --- Группировка по наименованию элемента (L3, без материала) ---
                add_name_groups(mssk_group, indices, 3)

            if mssk_group['children']:
                part_group['children'].append(mssk_group)

        result.append(part_group)

    return result


# ========== Excel Export ==========
def create_excel_report(groups: List[Dict], headers: List[str], rows: List[Dict], output_path: str) -> str:
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = "Группировка"
    
    columns = [
        ('Уровень', 8), ('Название группы', 60), ('Кол-во элементов', 15),
        ('Общий объём, м³', 18), ('Индексы элементов (№ п/п)', 40), ('Первый элемент', 12),
    ]
    
    for header in headers:
        columns.append((header, 20))
    
    area_columns = set()
    def collect_areas(groups_list):
        for g in groups_list:
            if g.get('total_areas'):
                area_columns.update(g['total_areas'].keys())
            if g.get('children'):
                collect_areas(g['children'])
    collect_areas(groups)
    
    for area_name in sorted(area_columns):
        columns.append((f'Суммарно: {area_name}', 20))
    
    for col_idx, (title, width) in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=title)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    row_num = 2
    
    def format_indices(indices):
        if not indices:
            return ''
        return ', '.join(str(i + 1) for i in sorted(indices))
    
    def write_group(group, level, indent=''):
        nonlocal row_num
        
        ws.cell(row=row_num, column=1, value=level)
        ws.cell(row=row_num, column=2, value=f"{indent}{group['name']}")
        ws.cell(row=row_num, column=3, value=group.get('count', 0))
        cell = ws.cell(row=row_num, column=4, value=group.get('total_volume', 0))
        cell.number_format = '#,##0.000'
        ws.cell(row=row_num, column=5, value=format_indices(group.get('indices', [])))
        
        indices = group.get('indices', [])
        if indices:
            ws.cell(row=row_num, column=6, value=sorted(indices)[0] + 1)
        else:
            ws.cell(row=row_num, column=6, value='')
        
        first = group.get('first_element', {})
        for col_idx, header in enumerate(headers, 7):
            ws.cell(row=row_num, column=col_idx, value=first.get(header, ''))
        
        area_start_col = 7 + len(headers)
        for i, area_name in enumerate(sorted(area_columns)):
            cell = ws.cell(row=row_num, column=area_start_col + i, 
                          value=group.get('total_areas', {}).get(area_name, 0))
            cell.number_format = '#,##0.00'
        
        row_num += 1
        
        for child in group.get('children', []):
            write_group(child, level + 1, indent + '  ')
    
    for group in groups:
        write_group(group, 0)
    
    ws2 = wb.create_sheet("Детали")
    
    detail_headers = ['№ группы', 'Название группы', 'Уровень'] + headers
    for col_idx, title in enumerate(detail_headers, 1):
        ws2.cell(row=1, column=col_idx, value=title)
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20
    
    detail_row = 2
    group_counter = [0]
    
    def write_detail_groups(groups_list, parent_name=''):
        nonlocal detail_row
        
        for group in groups_list:
            group_counter[0] += 1
            group_name = f"{parent_name} > {group['name']}" if parent_name else group['name']
            
            for idx in group.get('indices', []):
                if idx < len(rows):
                    row_data = rows[idx]
                    ws2.cell(row=detail_row, column=1, value=group_counter[0])
                    ws2.cell(row=detail_row, column=2, value=group_name)
                    ws2.cell(row=detail_row, column=3, value=group.get('level', 0))
                    
                    for col_idx, header in enumerate(headers, 4):
                        ws2.cell(row=detail_row, column=col_idx, value=row_data.get(header, ''))
                    
                    detail_row += 1
            
            if group.get('children'):
                write_detail_groups(group['children'], group_name)
    
    write_detail_groups(groups)
    
    wb.save(output_path)
    return output_path


# ========== Кеширование на диске ==========

def _get_cached_or_compute(input_excel_path: str, output_dir: str, 
                           processing_type: str, group_func, suffix: str) -> Dict[str, str]:
    """
    Проверяет, существует ли уже результат группировки на диске.
    Если да и входной файл не изменился — возвращает готовые пути.
    Если нет — вычисляет группировку.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    input_path = Path(input_excel_path)
    base_name = input_path.stem
    excel_output = output_dir / f"{base_name}_grouped{suffix}.xlsx"
    json_output = output_dir / f"{base_name}_grouped{suffix}.json"
    
    if os.path.exists(str(excel_output)) and os.path.exists(str(json_output)):
        input_mtime = os.path.getmtime(str(input_path))
        output_mtime = os.path.getmtime(str(excel_output))
        
        if output_mtime >= input_mtime:
            logger.info(f"Группировка ({processing_type}) уже существует, пропускаем "
                       f"({os.path.getsize(str(excel_output)) / 1024:.0f} KB)")
            return {
                'excel': str(excel_output),
                'json': str(json_output)
            }
        else:
            logger.info(f"Входной файл изменился, пересчитываем группировку ({processing_type})")
    
    logger.info(f"Вычисление группировки ({processing_type})...")
    start_time = time.time()
    
    df = pd.read_excel(input_excel_path)
    headers = df.columns.tolist()
    rows = df.to_dict('records')
    
    groups = group_func(rows, headers)
    
    create_excel_report(groups, headers, rows, str(excel_output))
    
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(groups, f, ensure_ascii=False, indent=2, default=str)
    
    elapsed = time.time() - start_time
    logger.info(f"Группировка ({processing_type}) вычислена за {elapsed:.1f} сек "
               f"({os.path.getsize(str(excel_output)) / 1024:.0f} KB)")
    
    return {
        'excel': str(excel_output),
        'json': str(json_output)
    }


# ========== Main Functions ==========
def process_ifc_excel(input_excel_path: str, output_dir: str = None) -> Dict[str, str]:
    """Process IFC Excel file and create grouped output (для КР). С кешированием на диске."""
    input_path = Path(input_excel_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_excel_path}")
    
    output_dir = output_dir or str(input_path.parent)
    
    return _get_cached_or_compute(
        input_excel_path=input_excel_path,
        output_dir=output_dir,
        processing_type='KR',
        group_func=group_elements,
        suffix=''
    )


def process_ifc_excel_ar(input_excel_path: str, output_dir: str = None) -> Dict[str, str]:
    """Process IFC Excel file and create grouped output (для АР). С кешированием на диске."""
    input_path = Path(input_excel_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_excel_path}")
    
    output_dir = output_dir or str(input_path.parent)
    
    return _get_cached_or_compute(
        input_excel_path=input_excel_path,
        output_dir=output_dir,
        processing_type='AR',
        group_func=group_elements_ar,
        suffix='_AR'
    )


def process_ifc_excel_new(input_excel_path: str, output_dir: str = None) -> Dict[str, str]:
    """Process IFC Excel file и создать grouped output (для КР) с новой упрощённой логикой
    геометрической группировки (как в group_excel_new.py). С кешированием на диске."""
    input_path = Path(input_excel_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_excel_path}")
    
    output_dir = output_dir or str(input_path.parent)
    
    return _get_cached_or_compute(
        input_excel_path=input_excel_path,
        output_dir=output_dir,
        processing_type='KR',
        group_func=group_elements_new,
        suffix='_new'
    )


def process_ifc_excel_mssk(input_excel_path: str, output_dir: str = None) -> Dict[str, str]:
    """Process IFC Excel file и создать grouped output (для КР) с группировкой
    по кодам МССК: Часть здания → Код МССК → Геометрия → Материал/Бетон.
    С кешированием на диске."""
    input_path = Path(input_excel_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_excel_path}")
    
    output_dir = output_dir or str(input_path.parent)
    
    return _get_cached_or_compute(
        input_excel_path=input_excel_path,
        output_dir=output_dir,
        processing_type='KR',
        group_func=group_elements_mssk,
        suffix='_mssk'
    )


def process_ifc_excel_ar_new(input_excel_path: str, output_dir: str = None) -> Dict[str, str]:
    """Process IFC Excel file и создать grouped output (для АР) с новой логикой из group_excel_new.py.

    В новой версии group_elements_ar идентична старой (за исключением float()-фиксов,
    которые уже применены к process_ifc_excel_ar), поэтому переиспользуем её."""
    return process_ifc_excel_ar(input_excel_path, output_dir)