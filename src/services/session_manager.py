import json
import os
import shutil
import threading
import uuid
import re
from datetime import datetime
from typing import Dict, List, Optional, Any
from collections import Counter
import pandas as pd
from werkzeug.utils import secure_filename
from src.core.prompt_manager import PromptManager
from src.core.logger import setup_logger
from src.services.zero_step import zero_step
from src.services.first_etap import first_step
from src.services.second_etap import second_step
from src.services.third_etap import third_step
from src.services.fourth_etap import fourth_step
from src.services.pdf_processor import process_pdf
from src.services.serializer import _make_glb_file
from src.services.group_excel import process_ifc_excel, process_ifc_excel_ar

from openpyxl import load_workbook

logger = setup_logger(__name__)


class SessionManager:
    """Управление сессиями обработки IFC файлов"""
    
    def __init__(self, upload_folder: str, output_folder: str, sessions_file: str, perechen_xlsx: str = None, koefs_xlsx: str = None):
        self.upload_folder = os.path.abspath(upload_folder)
        self.output_folder = os.path.abspath(output_folder)
        self.sessions_file = sessions_file
        self.perechen_xlsx = perechen_xlsx
        self.koefs_xlsx = koefs_xlsx
        self._sessions: Dict[str, Dict[str, Any]] = {}
        self._state_lock = threading.RLock()
        self._load()
        
        os.makedirs(upload_folder, exist_ok=True)
        os.makedirs(output_folder, exist_ok=True)
        os.makedirs(os.path.dirname(sessions_file) or ".", exist_ok=True)
        
        # Инициализируем PromptManager
        self.prompt_manager = PromptManager()
        self.prompt_manager.load_all()
    
    # =====================================================================
    #  БАЗОВЫЕ МЕТОДЫ (load/save/update/get/delete)
    # =====================================================================
    
    def _load(self) -> None:
        if os.path.exists(self.sessions_file):
            try:
                with open(self.sessions_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        self._sessions = {
                            k: v for k, v in data.items() 
                            if isinstance(v, dict) and "session_id" in v
                        }
                    else:
                        self._sessions = {}
            except json.JSONDecodeError as e:
                logger.error(f"Ошибка парсинга JSON в файле сессий: {e}")
                backup_path = f"{self.sessions_file}.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                try:
                    os.rename(self.sessions_file, backup_path)
                except Exception:
                    pass
                self._sessions = {}
            except Exception as e:
                logger.error(f"Ошибка загрузки сессий: {e}")
                self._sessions = {}
    
    def _save(self) -> None:
        try:
            temp_file = f"{self.sessions_file}.tmp"
            with open(temp_file, "w", encoding="utf-8") as f:
                json.dump(self._sessions, f, ensure_ascii=False, indent=2, default=str)
            
            if os.name == 'nt':
                if os.path.exists(self.sessions_file):
                    os.remove(self.sessions_file)
                os.rename(temp_file, self.sessions_file)
            else:
                os.replace(temp_file, self.sessions_file)
                
        except Exception as e:
            logger.error(f"Ошибка сохранения sessions.json: {e}")
    
    def _update(self, session_id: str, **fields) -> None:
        with self._state_lock:
            if session_id in self._sessions:
                protected_fields = {'session_id', 'created_at', 'ifc_file_path', 'pdf_file_path'}
                fields = {k: v for k, v in fields.items() if k not in protected_fields or k not in self._sessions[session_id]}
                self._sessions[session_id].update(fields)
                self._save()
    
    def _update_progress(self, session_id: str, progress: int, message: str) -> None:
        progress = max(0, min(100, progress))
        with self._state_lock:
            if session_id in self._sessions:
                self._sessions[session_id]["progress"] = progress
                self._sessions[session_id]["progress_message"] = message
                self._save()
    
    def get(self, session_id: str) -> Optional[Dict[str, Any]]:
        if not session_id or not isinstance(session_id, str):
            return None
            
        with self._state_lock:
            s = self._sessions.get(session_id)
            if not s:
                return None
            s = dict(s)
            
            # Если есть runs, подставляем файлы текущего запуска
            current_run_id = s.get("current_run_id")
            if current_run_id:
                runs = s.get("runs", [])
                for run in runs:
                    if run.get("run_id") == current_run_id:
                        s["files"] = run.get("files", [])
                        break
            
            self._decorate_files(s)
            return s
    
    def list_sessions(self) -> List[Dict[str, Any]]:
        with self._state_lock:
            items = [dict(s) for s in self._sessions.values()]
        
        items.sort(key=lambda s: s.get("created_at", ""), reverse=True)
        
        for s in items:
            self._decorate_files(s)
        return items
    
    def delete(self, session_id: str) -> bool:
        if not session_id or not isinstance(session_id, str):
            return False
            
        with self._state_lock:
            s = self._sessions.pop(session_id, None)
            if not s:
                return False
            self._save()
        
        session_dir = os.path.join(self.output_folder, session_id)
        if os.path.isdir(session_dir):
            try:
                shutil.rmtree(session_dir, ignore_errors=True)
            except Exception as e:
                logger.error(f"Ошибка удаления директории сессии {session_id}: {e}")
        
        upload_session_dir = os.path.join(self.upload_folder, session_id)
        if os.path.isdir(upload_session_dir):
            try:
                shutil.rmtree(upload_session_dir, ignore_errors=True)
            except Exception:
                pass
                
        return True
    
    def _decorate_files(self, session: Dict[str, Any]) -> None:
        sid = session.get("session_id")
        
        current_run_id = session.get("current_run_id")
        if current_run_id:
            runs = session.get("runs", [])
            for run in runs:
                if run.get("run_id") == current_run_id:
                    for f in run.get("files", []):
                        if isinstance(f, dict):
                            f["download_url"] = f"/ifc-vor/api/session/{sid}/download/{f.get('filename', '')}"
        
        for f in session.get("files", []):
            if isinstance(f, dict):
                f["download_url"] = f"/ifc-vor/api/session/{sid}/download/{f.get('filename', '')}"
    
    def file_path(self, session_id: str, filename: str) -> Optional[str]:
        if not filename or '..' in filename or '/' in filename or '\\' in filename:
            return None
            
        s = self.get(session_id)
        if not s:
            return None
        
        current_run_id = s.get("current_run_id")
        if current_run_id:
            runs = s.get("runs", [])
            for run in runs:
                if run.get("run_id") == current_run_id:
                    for f in run.get("files", []):
                        if f.get("filename") == filename:
                            return f.get("path")
        
        for f in s.get("files", []):
            if f.get("filename") == filename:
                return f.get("path")
        return None
    
    def _prepare_original_dir(self, session_dir: str) -> str:
        """Перемещает исходные файлы в original/ поддиректорию"""
        original_dir = os.path.join(session_dir, 'original')
        os.makedirs(original_dir, exist_ok=True)
        
        patterns_to_move = [
            r'^original_.*\.(ifc|pdf)$',
            r'^ДЛЯ_СМЕТЧИКА_.*\.xlsx$',
            r'^IFC_ВСЕ_ДАННЫЕ_.*\.xlsx$',
            r'^.*\.glb$',
            r'^blueprint_painted.*\.png$',
            r'^materials_colors\.md$',
        ]
        
        for f in os.listdir(session_dir):
            fpath = os.path.join(session_dir, f)
            if not os.path.isfile(fpath):
                continue
            
            should_move = any(re.match(p, f) for p in patterns_to_move)
            
            if should_move:
                dst = os.path.join(original_dir, f)
                shutil.move(fpath, dst)
                logger.info(f"Перемещён в original/: {f}")
        
        return original_dir
    
    def _get_original_excel_path(self, session_id: str) -> Optional[str]:
        """Находит путь к исходному Excel файлу в original/ директории"""
        s = self.get(session_id)
        if not s:
            return None
        
        session_dir = os.path.join(self.output_folder, session_id)
        original_dir = os.path.join(session_dir, 'original')
        
        if not os.path.exists(original_dir):
            return s.get("excel_file_path")
        
        # Приоритет: исправленный (имеет лист 'Данные')
        for f in os.listdir(original_dir):
            if 'ДЛЯ_СМЕТЧИКА' in f and 'исправленный' in f and f.endswith('.xlsx'):
                return os.path.join(original_dir, f)
        
        # Затем любой ДЛЯ_СМЕТЧИКА (кроме сокращенного)
        for f in os.listdir(original_dir):
            if 'ДЛЯ_СМЕТЧИКА' in f and 'сокращенный' not in f.lower() and f.endswith('.xlsx'):
                return os.path.join(original_dir, f)
        
        # Затем любой Excel (кроме сокращенного)
        for f in os.listdir(original_dir):
            if 'сокращенный' not in f.lower() and f.endswith('.xlsx'):
                return os.path.join(original_dir, f)
        
        # Last resort
        for f in os.listdir(original_dir):
            if f.endswith('.xlsx'):
                return os.path.join(original_dir, f)
        
        return s.get("excel_file_path")
    
    @staticmethod
    def _read_reference_json(path: str) -> List[Dict[str, Any]]:
        if not os.path.isfile(path):
            raise RuntimeError(f"Не сформирован файл {os.path.basename(path)}")
        with open(path, "r", encoding="utf-8") as stream:
            data = json.load(stream)
        if not isinstance(data, list) or any(not isinstance(item, dict) for item in data):
            raise RuntimeError(f"Файл {os.path.basename(path)} должен содержать массив объектов")
        return data
    
    # =====================================================================
    #  JSON-СПРАВОЧНИКИ (reference)
    # =====================================================================
    
    def build_reference(self, file, original_name: str, processing_type: str = "KR") -> Dict[str, Any]:
        """Запуск построения JSON-справочников из IFC или PDF."""
        if not file or not original_name:
            raise ValueError("Отсутствует файл или имя файла")

        extension = os.path.splitext(original_name)[1].lower()
        if extension not in (".ifc", ".pdf"):
            raise ValueError("Поддерживаются файлы .ifc и .pdf")

        processing_type = processing_type.upper()
        if processing_type not in ("KR", "AR"):
            processing_type = "KR"

        source_type = extension[1:]
        safe_name = secure_filename(original_name) or f"uploaded_file{extension}"
        session_id = str(uuid.uuid4())
        session_dir = os.path.join(self.output_folder, session_id)
        os.makedirs(session_dir, exist_ok=True)

        source_path = os.path.join(session_dir, f"original_{safe_name}")
        try:
            file.save(source_path)
            if not os.path.isfile(source_path) or os.path.getsize(source_path) == 0:
                raise ValueError("Ошибка сохранения файла")
        except Exception:
            logger.error("Ошибка сохранения файла для справочника", exc_info=True)
            raise

        session = {
            "session_id": session_id,
            "created_at": datetime.utcnow().isoformat() + "Z",
            "source_type": source_type,
            "status": "reference_processing",
            "processing_type": processing_type,
            "ifc_file_name": original_name if source_type == "ifc" else None,
            "ifc_file_path": source_path if source_type == "ifc" else None,
            "pdf_file_name": original_name if source_type == "pdf" else None,
            "pdf_file_path": source_path if source_type == "pdf" else None,
            "excel_file_name": None,
            "excel_file_path": None,
            "selected_rows": None,
            "construction_types": {},
            "construction_materials": {},
            "grouped_data": {},
            "building_height": None,
            "files": [],
            "runs": [],
            "current_run_id": None,
            "error": None,
            "progress": 10,
            "progress_message": "Формирование JSON-справочников...",
            "has_results": False,
            "is_reference_session": True,
        }

        with self._state_lock:
            self._sessions[session_id] = session
            self._save()

        thread = threading.Thread(
            target=self._build_reference_bg,
            args=(session_id, source_path, source_type, processing_type),
            daemon=True,
            name=f"Reference-{source_type.upper()}-{session_id[:8]}",
        )
        try:
            thread.start()
        except Exception as exc:
            self._update(
                session_id,
                status="error",
                error=f"{type(exc).__name__}: {exc}",
                progress_message="Не удалось запустить фоновую обработку",
            )
            raise

        return {
            "session_id": session_id,
            "source_type": source_type,
            "status": "reference_processing",
            "message": "Файл принят, формирование JSON-справочников запущено",
        }

    def _build_reference_bg(self, session_id: str, source_path: str, source_type: str, processing_type: str = "KR") -> None:
        try:
            session_dir = os.path.join(self.output_folder, session_id)
            self._update_progress(session_id, 20, "Извлечение и группировка элементов...")

            if source_type == "ifc":
                from src.services.ifc_reference_builder import build_reference_from_ifc
                build_reference_from_ifc(source_path, session_dir, processing_type)
            else:
                process_pdf(
                    source_path,
                    output_folder=session_dir,
                    reference_only=True,
                    processing_type=processing_type,
                )

            elements_path = os.path.join(session_dir, "ifc_elements_output.json")
            grouped_path = os.path.join(session_dir, "ifc_raw_elements_grouped.json")
            self._read_reference_json(elements_path)
            self._read_reference_json(grouped_path)

            files = [
                {"path": elements_path, "filename": "ifc_elements_output.json", "size": os.path.getsize(elements_path)},
                {"path": grouped_path, "filename": "ifc_raw_elements_grouped.json", "size": os.path.getsize(grouped_path)},
            ]

            with self._state_lock:
                current = self._sessions.get(session_id)
                if current is not None:
                    current["status"] = "completed"
                    current["files"] = files
                    current["progress"] = 100
                    current["progress_message"] = "JSON-справочники сформированы"
                    current["has_results"] = True
                    self._save()
        except Exception as exc:
            error_msg = f"{type(exc).__name__}: {exc}"
            logger.error(f"Ошибка построения JSON-справочников для сессии {session_id}", exc_info=True)
            self._update(
                session_id,
                status="error",
                error=error_msg,
                progress_message="Ошибка формирования JSON-справочников",
            )

    def get_reference_result(self, session_id: str) -> Optional[Dict[str, Any]]:
        session = self.get(session_id)
        if not session:
            raise KeyError("Сессия не найдена")
        if not session.get("is_reference_session"):
            raise ValueError("Сессия не относится к построению справочников")
        if session.get("status") == "reference_processing":
            return None
        if session.get("status") == "error":
            raise RuntimeError(session.get("error") or "Ошибка построения справочников")
        if session.get("status") != "completed":
            return None

        session_dir = os.path.join(self.output_folder, session_id)
        return {
            "session_id": session_id,
            "source_type": session.get("source_type", ""),
            "ifc_elements_output": self._read_reference_json(
                os.path.join(session_dir, "ifc_elements_output.json")
            ),
            "ifc_raw_elements_grouped": self._read_reference_json(
                os.path.join(session_dir, "ifc_raw_elements_grouped.json")
            ),
        }

    # =====================================================================
    #  ОБРАБОТКА IFC
    # =====================================================================
    
    def process_ifc(self, file, original_name: str, processing_type: str = "KR") -> Dict[str, Any]:
        if not file or not original_name:
            raise ValueError("Отсутствует файл или имя файла")
        
        processing_type = processing_type.upper()
        if processing_type not in ("KR", "AR"):
            processing_type = "KR"
        
        safe_name = secure_filename(original_name)
        if not safe_name:
            safe_name = "uploaded_file.ifc"
        
        session_id = str(uuid.uuid4())
        session_dir = os.path.join(self.output_folder, session_id)
        os.makedirs(session_dir, exist_ok=True)
        
        ifc_filename = f"original_{safe_name}"
        ifc_path = os.path.join(session_dir, ifc_filename)
        
        try:
            file.save(ifc_path)
            if not os.path.exists(ifc_path) or os.path.getsize(ifc_path) == 0:
                raise ValueError("Ошибка сохранения файла")
        except Exception as e:
            logger.error(f"Ошибка сохранения IFC файла: {e}")
            raise
        
        session = {
            "session_id": session_id,
            "created_at": datetime.utcnow().isoformat() + "Z",
            "source_type": "ifc",
            "status": "ifc_processing",
            "processing_type": processing_type,
            "ifc_file_name": original_name,
            "ifc_file_path": ifc_path,
            "excel_file_name": None,
            "excel_file_path": None,
            "selected_rows": None,
            "construction_types": {},
            "construction_materials": {},
            "grouped_data": {},
            "building_height": None,
            "files": [],
            "runs": [],
            "current_run_id": None,
            "error": None,
            "progress": 0,
            "progress_message": f"Начало обработки IFC ({processing_type})...",
            "has_results": False,
        }
        
        with self._state_lock:
            self._sessions[session_id] = session
            self._save()
        
        thread = threading.Thread(
            target=self._process_ifc_bg,
            args=(session_id, ifc_path, processing_type),
            daemon=True,
            name=f"IFC-Processing-{session_id[:8]}"
        )
        thread.start()
        
        return {
            "session_id": session_id,
            "status": "ifc_processing",
            "message": f"IFC файл принят, начата обработка ({processing_type})",
        }
    
    def _process_ifc_bg(self, session_id: str, ifc_path: str, processing_type: str = "KR") -> None:
        try:
            self._update_progress(session_id, 5, "Проверка IFC файла...")
            
            if not os.path.exists(ifc_path):
                raise FileNotFoundError(f"IFC файл не найден: {ifc_path}")
            
            session_dir = os.path.join(self.output_folder, session_id)
            
            self._update_progress(session_id, 10, f"Обработка IFC файла ({processing_type})...")
            self._update_progress(session_id, 20, "Извлечение элементов из IFC...")
            zero_step(ifc_path, output_folder=session_dir)

            # Формируем справочные JSON
            try:
                from src.services.ifc_reference_builder import build_reference_from_ifc
                build_reference_from_ifc(ifc_path, session_dir, processing_type)
            except Exception as e:
                logger.warning(f"Не удалось сформировать JSON-файлы справочника для IFC: {e}", exc_info=True)

            self._update_progress(session_id, 80, "Проверка результатов...")
            
            excel_for_smetchik = os.path.join(session_dir, 'ДЛЯ_СМЕТЧИКА_исправленный.xlsx')
            
            if not os.path.exists(excel_for_smetchik):
                excel_files = [f for f in os.listdir(session_dir) if f.endswith(('.xlsx', '.xls'))]
                if excel_files:
                    excel_for_smetchik = os.path.join(session_dir, excel_files[0])
                else:
                    raise RuntimeError("Не удалось найти созданный Excel файл")
            
            # Создаём GLB модель
            try:
                _make_glb_file(ifc_path, session_dir)
            except Exception as e:
                logger.warning(f'Не удалось создать файл 3D модели: {e}')
            
            # Перемещаем исходные файлы в original/
            original_dir = self._prepare_original_dir(session_dir)
            
            # Находим Excel в original/
            excel_path = None
            for f in os.listdir(original_dir):
                if 'ДЛЯ_СМЕТЧИКА' in f and f.endswith('.xlsx'):
                    excel_path = os.path.join(original_dir, f)
                    excel_filename = f
                    break
            
            if not excel_path:
                for f in os.listdir(original_dir):
                    if f.endswith('.xlsx'):
                        excel_path = os.path.join(original_dir, f)
                        excel_filename = f
                        break
            
            if not excel_path:
                raise RuntimeError("Не удалось сохранить Excel файл")
            
            with self._state_lock:
                if session_id in self._sessions:
                    self._sessions[session_id]["excel_file_name"] = excel_filename
                    self._sessions[session_id]["excel_file_path"] = excel_path
                    self._sessions[session_id]["status"] = "ifc_processed"
                    self._sessions[session_id]["progress"] = 100
                    self._sessions[session_id]["progress_message"] = f"Обработка завершена ({processing_type}). Выберите строки."
                    self._save()
            
        except Exception as e:
            import traceback
            error_msg = f"{type(e).__name__}: {str(e)}"
            logger.error(f"Ошибка обработки IFC для сессии {session_id}:\n{traceback.format_exc()}")
            self._update(session_id, status="error", error=error_msg)
    
    # =====================================================================
    #  ОБРАБОТКА PDF
    # =====================================================================
    
    def process_pdf(self, file, original_name: str, processing_type: str = "KR") -> Dict[str, Any]:
        if not file or not original_name:
            raise ValueError("Отсутствует файл или имя файла")
        
        processing_type = processing_type.upper()
        if processing_type not in ("KR", "AR"):
            processing_type = "KR"
        
        safe_name = secure_filename(original_name)
        if not safe_name:
            safe_name = "uploaded_file.pdf"
        
        session_id = str(uuid.uuid4())
        session_dir = os.path.join(self.output_folder, session_id)
        os.makedirs(session_dir, exist_ok=True)
        
        pdf_filename = f"original_{safe_name}"
        pdf_path = os.path.join(session_dir, pdf_filename)
        
        try:
            file.save(pdf_path)
            if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
                raise ValueError("Ошибка сохранения файла")
        except Exception as e:
            logger.error(f"Ошибка сохранения PDF файла: {e}")
            raise
        
        session = {
            "session_id": session_id,
            "created_at": datetime.utcnow().isoformat() + "Z",
            "source_type": "pdf",
            "status": "pdf_processing",
            "processing_type": processing_type,
            "pdf_file_name": original_name,
            "pdf_file_path": pdf_path,
            "ifc_file_name": None,
            "ifc_file_path": None,
            "excel_file_name": None,
            "excel_file_path": None,
            "selected_rows": None,
            "construction_types": {},
            "construction_materials": {},
            "grouped_data": {},
            "building_height": None,
            "files": [],
            "runs": [],
            "current_run_id": None,
            "error": None,
            "progress": 0,
            "progress_message": f"Начало обработки PDF ({processing_type})...",
            "has_results": False,
        }
    
        with self._state_lock:
            self._sessions[session_id] = session
            self._save()
        
        thread = threading.Thread(
            target=self._process_pdf_bg,
            args=(session_id, pdf_path, processing_type),
            daemon=True,
            name=f"PDF-Processing-{session_id[:8]}"
        )
        thread.start()
        
        return {
            "session_id": session_id,
            "status": "pdf_processing",
            "message": f"PDF файл принят, начата обработка ({processing_type})",
        }
    
    def _process_pdf_bg(self, session_id: str, pdf_path: str, processing_type: str = "KR") -> None:
        try:
            if not os.path.exists(pdf_path):
                raise FileNotFoundError(f"PDF файл не найден: {pdf_path}")
            
            session_dir = os.path.join(self.output_folder, session_id)
            
            self._update_progress(session_id, 20, "Извлечение элементов из чертежа...")
            result = self._process_pdf_with_progress(session_id, pdf_path, session_dir)
            
            self._update_progress(session_id, 90, "Проверка результатов...")
            
            excel_for_smetchik = result["excel_smetchik_path"]
            excel_all_data = result["excel_all_data_path"]
            
            excel_for_smetchik, excel_all_data = self._check_and_merge_sheets(
                excel_for_smetchik, excel_all_data
            )
            
            if not os.path.exists(excel_for_smetchik):
                raise RuntimeError("Не удалось найти созданный Excel файл")
            
            original_dir = self._prepare_original_dir(session_dir)
            
            excel_path = None
            for f in os.listdir(original_dir):
                if 'ДЛЯ_СМЕТЧИКА' in f and f.endswith('.xlsx'):
                    excel_path = os.path.join(original_dir, f)
                    excel_filename = f
                    break
            
            if not excel_path:
                for f in os.listdir(original_dir):
                    if f.endswith('.xlsx'):
                        excel_path = os.path.join(original_dir, f)
                        excel_filename = f
                        break
            
            if not excel_path:
                raise RuntimeError("Не удалось найти Excel файл в original/")
            
            with self._state_lock:
                if session_id in self._sessions:
                    self._sessions[session_id]["excel_file_name"] = excel_filename
                    self._sessions[session_id]["excel_file_path"] = excel_path
                    self._sessions[session_id]["status"] = "ifc_processed"
                    self._sessions[session_id]["progress"] = 100
                    self._sessions[session_id]["progress_message"] = f"Обработка завершена ({processing_type}). Выберите строки."
                    self._save()
                    
        except Exception as e:
            import traceback
            error_msg = f"{type(e).__name__}: {str(e)}"
            logger.error(f"Ошибка обработки PDF для сессии {session_id}:\n{traceback.format_exc()}")
            self._update(session_id, status="error", error=error_msg)
    
    def _check_and_merge_sheets(self, excel_for_smetchik, excel_all_data):
        def get_numbered_sheets(filepath):
            wb = load_workbook(filepath, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            pattern = re.compile(r'^Данные_\d+$')
            return [name for name in sheet_names if pattern.match(name)]
        
        def merge_sheets_to_file(source_filepath, output_filename):
            numbered_sheets = get_numbered_sheets(source_filepath)
            if not numbered_sheets:
                return source_filepath
            all_data = []
            for sheet_name in numbered_sheets:
                df = pd.read_excel(source_filepath, sheet_name=sheet_name)
                df['Источник_лист'] = sheet_name
                all_data.append(df)
            merged_df = pd.concat(all_data, ignore_index=True)
            output_path = os.path.join(os.path.dirname(source_filepath), output_filename)
            merged_df.to_excel(output_path, sheet_name='Данные', index=False)
            return output_path
        
        return (
            merge_sheets_to_file(excel_for_smetchik, 'ДЛЯ_СМЕТЧИКА_объединенный.xlsx'),
            merge_sheets_to_file(excel_all_data, 'IFC_ВСЕ_ДАННЫЕ_объединенный.xlsx'),
        )
    
    def _process_pdf_with_progress(self, session_id: str, pdf_path: str, session_dir: str) -> Dict[str, str]:
        last_update_time = [0]
        min_interval = 2.0
        
        def progress_callback(stage_name: str, progress_percent: int):
            import time
            current_time = time.time()
            if current_time - last_update_time[0] < min_interval and progress_percent < 100:
                return
            last_update_time[0] = current_time
            self._update_progress(session_id, progress_percent, stage_name)
        
        result = process_pdf(pdf_path, output_folder=session_dir, progress_callback=progress_callback)
        self._update_progress(session_id, 90, "Проверка результатов...")
        return result
    
    # =====================================================================
    #  ЗАПУСКИ (runs)
    # =====================================================================
    
    def new_run(self, session_id: str, row_indices: List[int], 
                construction_types: Dict[int, str] = None,
                construction_materials: Dict[int, str] = None,
                building_height: float = None, 
                grouped_data: Dict[str, Any] = None,
                processing_type: str = "KR") -> Dict[str, Any]:
        
        s = self.get(session_id)
        if not s:
            raise KeyError("Сессия не найдена")
        
        processing_type = processing_type.upper()
        if processing_type not in ("KR", "AR"):
            processing_type = "KR"
        
        if s["status"] not in ("ifc_processed", "selecting_rows", "completed"):
            raise RuntimeError(f"Неверный статус сессии: {s['status']}")
        
        if not row_indices:
            raise ValueError("Необходимо выбрать хотя бы одну строку")
        
        row_indices = [int(i) for i in row_indices if isinstance(i, (int, float)) and i >= 0]
        if not row_indices:
            raise ValueError("Некорректные индексы строк")
        
        excel_path = self._get_original_excel_path(session_id)
        
        if not excel_path or not os.path.exists(excel_path):
            logger.warning(f"original/ не найден, использую excel_file_path из сессии")
            excel_path = s.get("excel_file_path")
        
        if not excel_path or not os.path.exists(excel_path):
            raise RuntimeError(f"Исходный Excel файл не найден")
        
        logger.info(f"Новый запуск: сессия={session_id}, Excel={excel_path}, строк={len(row_indices)}, тип={processing_type}")
        
        run_id = str(uuid.uuid4())
        runs = s.get('runs', [])
        run_number = len(runs) + 1
        session_dir = os.path.join(self.output_folder, session_id)
        run_dir = os.path.join(session_dir, f'run_{run_number:03d}')
        os.makedirs(run_dir, exist_ok=True)
        
        run_excel_name = os.path.basename(excel_path)
        run_excel_path = os.path.join(run_dir, run_excel_name)
        shutil.copy2(excel_path, run_excel_path)
        
        # Копируем IFC_ВСЕ_ДАННЫЕ и GLB
        original_dir = os.path.join(session_dir, 'original')
        search_dir = original_dir if os.path.exists(original_dir) else session_dir
        
        for f in os.listdir(search_dir):
            if 'IFC_ВСЕ_ДАННЫЕ' in f and f.endswith('.xlsx'):
                shutil.copy2(os.path.join(search_dir, f), os.path.join(run_dir, f))
                break
        
        for f in os.listdir(search_dir):
            if f.endswith('.glb'):
                shutil.copy2(os.path.join(search_dir, f), os.path.join(run_dir, f))
                break
        
        for f in os.listdir(search_dir):
            if 'сокращенный' in f.lower() and f.endswith('.xlsx'):
                shutil.copy2(os.path.join(search_dir, f), os.path.join(run_dir, f))
                break
        
        run = {
            "run_id": run_id,
            "run_number": run_number,
            "processing_type": processing_type,
            "status": "processing",
            "selected_rows": row_indices,
            "construction_types": construction_types or {},
            "construction_materials": construction_materials or {},
            "building_height": building_height,
            "grouped_data": grouped_data or {},
            "files": [],
            "created_at": datetime.utcnow().isoformat() + "Z",
        }
        
        runs.append(run)
        
        self._update(
            session_id,
            current_run_id=run_id,
            runs=runs,
            status="processing",
            progress=0,
            progress_message=f"Запуск {run_number} ({processing_type}): выбрано {len(row_indices)} строк"
        )
        
        thread = threading.Thread(
            target=self._run_processing_pipeline_in_run,
            args=(session_id, run_id, run_number, run_dir, run_excel_path, 
                row_indices, construction_types or {}, construction_materials or {}, 
                building_height, processing_type),
            daemon=True,
            name=f"Pipeline-Run{run_number}-{session_id[:8]}"
        )
        thread.start()
        
        return {
            "session_id": session_id,
            "run_id": run_id,
            "run_number": run_number,
            "processing_type": processing_type,
            "status": "processing",
            "selected_rows": len(row_indices),
            "message": f"Запуск {run_number} ({processing_type}): выбрано {len(row_indices)} строк, начата обработка"
        }
    
    def switch_run(self, session_id: str, run_id: str) -> Dict[str, Any]:
        s = self.get(session_id)
        if not s:
            raise KeyError("Сессия не найдена")
        
        runs = s.get('runs', [])
        target_run = next((r for r in runs if r.get('run_id') == run_id), None)
        
        if not target_run:
            raise ValueError(f"Запуск {run_id} не найден")
        
        self._update(session_id, current_run_id=run_id)
        
        return {
            "session_id": session_id,
            "run_id": run_id,
            "run_number": target_run.get('run_number'),
            "processing_type": target_run.get('processing_type', 'KR'),
            "status": target_run.get('status'),
            "files": target_run.get('files', []),
            "building_height": target_run.get('building_height'),
        }
    
    def get_run(self, session_id: str, run_id: str) -> Optional[Dict[str, Any]]:
        s = self.get(session_id)
        if not s:
            return None
        runs = s.get('runs', [])
        for run in runs:
            if run.get('run_id') == run_id:
                return dict(run)
        return None
    
    def list_runs(self, session_id: str) -> List[Dict[str, Any]]:
        s = self.get(session_id)
        if not s:
            return []
        return s.get('runs', [])
    
    def select_rows(self, session_id: str, row_indices: List[int], 
                    all_rows: bool = False, row_types: Dict[int, str] = None,
                    row_materials: Dict[int, str] = None,
                    building_height: float = None, grouped_data: Dict[str, Any] = None,
                    processing_type: str = "KR") -> Dict[str, Any]:
        s = self.get(session_id)
        if not s:
            raise KeyError("Сессия не найдена")
        
        if s["status"] not in ("ifc_processed", "selecting_rows"):
            raise RuntimeError(f"Неверный статус сессии: {s['status']}")
        
        if not all_rows and not row_indices:
            raise ValueError("Необходимо выбрать хотя бы одну строку")
        
        if not all_rows:
            row_indices = [int(i) for i in row_indices if isinstance(i, (int, float)) and i >= 0]
            if not row_indices:
                raise ValueError("Некорректные индексы строк")
        
        if all_rows:
            excel_path = self._get_original_excel_path(session_id)
            if not excel_path or not os.path.exists(excel_path):
                raise RuntimeError("Excel файл не найден")
            try:
                df = pd.read_excel(excel_path)
                row_indices = list(range(len(df)))
            except Exception as e:
                raise RuntimeError(f"Ошибка чтения Excel файла: {str(e)}")
        
        return self.new_run(
            session_id, row_indices, 
            row_types or {}, row_materials or {},
            building_height, grouped_data or {},
            processing_type
        )
    
    def _run_processing_pipeline_in_run(self, session_id: str, run_id: str, 
                                         run_number: int, run_dir: str,
                                         excel_path: str, row_indices: List[int],
                                         construction_types: Dict[int, str],
                                         construction_materials: Dict[int, str],
                                         building_height: float = None,
                                         processing_type: str = "KR") -> None:
        try:
            processing_type = processing_type.upper()
            if processing_type not in ("KR", "AR"):
                processing_type = "KR"
            
            logger.info(f"Запуск пайплайна: run={run_number}, тип={processing_type}")
            
            # Применяем материалы
            if construction_materials:
                try:
                    wb = load_workbook(excel_path)
                    ws = wb['Данные']
                    
                    material_col = None
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value == 'Материал':
                            material_col = col_idx
                            break
                    
                    if material_col:
                        updated_count = 0
                        for row_idx in range(2, ws.max_row + 1):
                            data_idx = row_idx - 2
                            if str(data_idx) in construction_materials:
                                material = construction_materials[str(data_idx)]
                                if material and material != '-':
                                    ws.cell(row=row_idx, column=material_col).value = material
                                    updated_count += 1
                        
                        if updated_count > 0:
                            wb.save(excel_path)
                            logger.info(f"Excel обновлён: {updated_count} материалов изменено")
                    
                    wb.close()
                    
                    with open(os.path.join(run_dir, 'materials.json'), 'w', encoding='utf-8') as f:
                        json.dump(construction_materials, f, ensure_ascii=False, indent=2)
                except Exception as e:
                    logger.error(f"Ошибка при применении материалов к Excel: {e}")
            
            # Фильтрация
            self._update_progress(session_id, 5, f"Запуск {run_number}: Фильтрация элементов...")
            
            df_original = pd.read_excel(excel_path, sheet_name='Данные')
            unique_indices = sorted(set(row_indices))
            df_filtered = df_original.iloc[unique_indices].reset_index(drop=True)
            
            filtered_path = os.path.join(run_dir, 'filtered_elements.xlsx')
            with pd.ExcelWriter(filtered_path, engine='openpyxl') as writer:
                df_filtered.to_excel(writer, sheet_name='Данные', index=False)
            
            logger.info(f"Отфильтровано {len(df_filtered)} элементов из {len(df_original)}")
            
            # Группировка (разная для КР и АР)
            self._update_progress(session_id, 10, f"Запуск {run_number}: Группировка элементов ({processing_type})...")
            
            if processing_type == "AR":
                group_result = process_ifc_excel_ar(filtered_path, run_dir)
            else:
                group_result = process_ifc_excel(filtered_path, run_dir)
            
            # Кешируем дерево проекта в original/
            session_dir = os.path.join(self.output_folder, session_id)
            original_dir = os.path.join(session_dir, 'original')
            all_data_path = None
            
            if os.path.exists(original_dir):
                for f in os.listdir(original_dir):
                    if 'IFC_ВСЕ_ДАННЫЕ' in f and f.endswith('.xlsx'):
                        all_data_path = os.path.join(original_dir, f)
                        break
            
            if all_data_path and os.path.exists(all_data_path):
                os.makedirs(original_dir, exist_ok=True)
                
                if processing_type == "AR":
                    all_project_tree = process_ifc_excel_ar(all_data_path, original_dir)
                else:
                    all_project_tree = process_ifc_excel(all_data_path, original_dir)
                
                whole_tree_src = all_project_tree['excel']
                whole_tree_dst = os.path.join(run_dir, 'Дерево_проекта.xlsx')
                if os.path.exists(whole_tree_src):
                    shutil.copy2(whole_tree_src, whole_tree_dst)
                    logger.info(f"Дерево проекта скопировано из кеша в {whole_tree_dst}")
            
            tree_excel_src = group_result['excel']
            tree_excel_dst = os.path.join(run_dir, 'Дерево_проекта_выбранные_элементы.xlsx')
            if os.path.exists(tree_excel_src) and tree_excel_src != tree_excel_dst:
                if os.path.exists(tree_excel_dst):
                    os.remove(tree_excel_dst)
                os.rename(tree_excel_src, tree_excel_dst)
            
            # Загружаем JSON с группами
            with open(group_result['json'], 'r', encoding='utf-8') as f:
                groups = json.load(f)
            
            # Собираем листовые группы
            self._update_progress(session_id, 15, f"Запуск {run_number}: Формирование групп для сметчика...")
            
            def collect_leaf_groups(groups_list, result=None):
                if result is None:
                    result = []
                for group in groups_list:
                    if group.get('children') and len(group['children']) > 0:
                        collect_leaf_groups(group['children'], result)
                    else:
                        result.append(group)
                return result
            
            leaf_groups = collect_leaf_groups(groups)
            
            smetchik_rows = []
            for group in leaf_groups:
                first_element = dict(group.get('first_element', {}))
                row_data = first_element.copy()
                
                row_data['Объём_NetVolume_м3_grouped'] = group.get('total_volume', 0)
                row_data['Количество_в_группе_grouped'] = group.get('count', 1)
                
                for area_name, area_value in group.get('total_areas', {}).items():
                    key = area_name if area_name.endswith('_grouped') else f'{area_name}_grouped'
                    row_data[key] = area_value
                
                row_data['Название_группы'] = group.get('name', '')
                row_data['Уровень_группы'] = group.get('level', 0)
                row_data['Индексы_элементов'] = ', '.join(str(i + 1) for i in group.get('indices', []))
                
                smetchik_rows.append(row_data)
            
            df_smetchik = pd.DataFrame(smetchik_rows)
            smetchik_path = os.path.join(run_dir, 'ДЛЯ_СМЕТЧИКА_сгруппированный.xlsx')
            
            grouped_cols = [c for c in df_smetchik.columns if c.endswith('_grouped')]
            info_cols = ['Название_группы', 'Уровень_группы', 'Индексы_элементов']
            other_cols = [c for c in df_smetchik.columns if c not in grouped_cols and c not in info_cols]
            
            for col in info_cols + grouped_cols:
                if col not in df_smetchik.columns:
                    df_smetchik[col] = ''
            
            existing_other = [c for c in other_cols if c in df_smetchik.columns]
            existing_grouped = [c for c in grouped_cols if c in df_smetchik.columns]
            existing_info = [c for c in info_cols if c in df_smetchik.columns]
            df_smetchik = df_smetchik[existing_other + existing_grouped + existing_info]
            
            with pd.ExcelWriter(smetchik_path, engine='openpyxl') as writer:
                df_smetchik.to_excel(writer, sheet_name='Данные', index=False)
            
            # Определяем часть здания для каждой группы
            new_construction_types = {}
            for i, group in enumerate(leaf_groups):
                indices = group.get('indices', [])
                parts_in_group = []
                for idx in indices:
                    part = construction_types.get(str(idx), construction_types.get(idx, None))
                    if part:
                        parts_in_group.append(part)
                
                if parts_in_group:
                    part_counts = Counter(parts_in_group)
                    new_construction_types[str(i)] = part_counts.most_common(1)[0][0]
                else:
                    new_construction_types[str(i)] = 'Надземная'
            
            with open(os.path.join(run_dir, 'building_parts.json'), 'w', encoding='utf-8') as f:
                json.dump(new_construction_types, f, ensure_ascii=False, indent=2)
            
            # Обновляем пути для этапов
            excel_path = smetchik_path
            row_indices = list(range(len(df_smetchik)))
            
            # Этапы 1-4
            self._update_progress(session_id, 20, f"Запуск {run_number}: Этап 1 — Анализ через LLM...")
            first_step(
                prompt_manager=self.prompt_manager,
                file=excel_path,
                rows=[i+1 for i in row_indices],
                output_folder=run_dir
            )
            
            self._update_progress(session_id, 40, f"Запуск {run_number}: Этап 2 — Фильтрация по части здания...")
            second_step(input_folder=run_dir)
            
            self._update_progress(session_id, 60, f"Запуск {run_number}: Этап 3 — Фильтрация по высоте...")
            third_step(input_folder=run_dir, building_height=building_height)
            
            self._update_progress(session_id, 90, f"Запуск {run_number}: Этап 4 — Формирование перечня ({processing_type})...")
            fourth_step(input_folder=run_dir, processing_type=processing_type)
            
            self._update_progress(session_id, 95, f"Запуск {run_number}: Сохранение результатов...")
            
            # Собираем финальные файлы
            final_files = []
            skip_patterns = {'filtered_elements.xlsx', 'building_parts.json', 'materials.json'}
            
            for f in os.listdir(run_dir):
                fpath = os.path.join(run_dir, f)
                if os.path.isfile(fpath):
                    if f in skip_patterns:
                        continue
                    if f.endswith('json'):
                        continue
                    if f.startswith('Нормализованные_данные_элемента_') or f.endswith('.ifc'):
                        continue
                    if any(f.startswith(p) for p in ['Промежуточные_работы_', 'height', 'Финальный', 'Подобранные', 'Все_найденные']):
                        continue
                    
                    final_files.append({
                        "path": fpath,
                        "filename": f,
                        "size": os.path.getsize(fpath)
                    })
            
            final_files.sort(key=lambda x: x['filename'])
            
            # Добавляем справочные JSON из корня сессии
            for src_name, display_name in [
                ("ifc_elements_output.json", "все элементы.json"),
                ("ifc_raw_elements_grouped.json", "группы элементов.json"),
                ("ifc_raw_elements_grouped.xlsx", "группы элементов.xlsx"),
            ]:
                src_path = os.path.join(session_dir, src_name)
                if os.path.isfile(src_path):
                    final_files.append({
                        "path": src_path,
                        "filename": display_name,
                        "size": os.path.getsize(src_path),
                    })
            
            # Обновляем run
            with self._state_lock:
                if session_id in self._sessions:
                    runs = self._sessions[session_id].get('runs', [])
                    for run in runs:
                        if run['run_id'] == run_id:
                            run['status'] = 'completed'
                            run['files'] = final_files
                            run['building_height'] = building_height
                            break
                    
                    self._sessions[session_id]['runs'] = runs
                    self._sessions[session_id]['status'] = 'completed'
                    self._sessions[session_id]['has_results'] = True
                    self._sessions[session_id]['progress'] = 100
                    self._sessions[session_id]['progress_message'] = f"Запуск {run_number} ({processing_type}) завершён"
                    self._save()
            
        except Exception as e:
            import traceback
            error_msg = f"{type(e).__name__}: {str(e)}"
            logger.error(f"Ошибка пайплайна для запуска {run_id}:\n{traceback.format_exc()}")
            
            with self._state_lock:
                if session_id in self._sessions:
                    runs = self._sessions[session_id].get('runs', [])
                    for run in runs:
                        if run['run_id'] == run_id:
                            run['status'] = 'error'
                            run['error'] = error_msg
                            break
                    self._sessions[session_id]['runs'] = runs
                    self._save()
    
    # =====================================================================
    #  ФИЛЬТРАЦИЯ ПО ВЫСОТЕ
    # =====================================================================
    
    def filter_by_height(self, session_id: str, building_height: float) -> Dict[str, Any]:
        s = self.get(session_id)
        if not s:
            raise KeyError("Сессия не найдена")
        
        if s["status"] not in ("filtering_height", "filtering_type", "processing"):
            raise RuntimeError(f"Неверный статус: {s['status']}")
        
        if not isinstance(building_height, (int, float)) or building_height <= 0:
            raise ValueError("Высота здания должна быть положительным числом")
        
        if building_height > 10000:
            raise ValueError("Слишком большая высота здания")
        
        self._update(session_id, building_height=building_height)
        
        return {
            "session_id": session_id,
            "status": "processing",
            "building_height": building_height,
            "message": f"Высота обновлена: {building_height}м"
        }