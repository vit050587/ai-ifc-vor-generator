from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import os
import copy
from datetime import datetime

DEFAULT_FLOOR_HEIGHT = 3.0
DEFAULT_WALL_THICKNESS = 380
DEFAULT_SLAB_THICKNESS = 200
DEFAULT_COLUMN_SIZE = 400

OUTPUT_DIR = "drawing_analysis"

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def calculate_volumes(data: dict) -> dict:
    """Рассчитывает объёмы с проверкой на None"""
    
     
    for wall in data.get('walls', []):
        thickness = safe_float(wall.get('thickness_mm'), DEFAULT_WALL_THICKNESS)
        length = safe_float(wall.get('length_m'), 0)
        height = DEFAULT_FLOOR_HEIGHT
        
        if length > 0:
            wall['area_m2'] = round(length * height, 2)
            wall['volume_m3'] = round(length * height * (thickness / 1000), 2)
            length_mm = length * 1000
            wall['perimeter_mm'] = round(
                2 * (length_mm + thickness),
                0,
            )
        else:
            wall['area_m2'] = 'не указана'
            wall['volume_m3'] = 'не указан'
    
     
    for col in data.get('columns', []):
        w = safe_float(col.get('width_mm'), DEFAULT_COLUMN_SIZE)
        h = safe_float(col.get('height_mm'), DEFAULT_COLUMN_SIZE)
        height_m = DEFAULT_FLOOR_HEIGHT
        col['volume_m3'] = round((w / 1000) * (h / 1000) * height_m, 2)
        col['section'] = f"{int(w)}x{int(h)}"
    
     
    for slab in data.get('slabs', []):
        area = safe_float(slab.get('area_m2'), 0)
        thickness = safe_float(slab.get('thickness_mm'), DEFAULT_SLAB_THICKNESS)
        if area > 0:
            slab['volume_m3'] = round(area * thickness / 1000, 2)
            slab['volume_liters'] = round(area * thickness, 2)
    
     
    for space in data.get('spaces', []):
        area = safe_float(space.get('area_m2'), 0)
        space['volume_m3'] = round(area * DEFAULT_FLOOR_HEIGHT, 2)
    
    return data

def form_wall_type_statistics(data: dict) -> dict:
    calculated_data = calculate_volumes(copy.deepcopy(data))
    statistics_by_type = {}

    for wall in calculated_data.get('walls', []):
        wall_type = wall.get('material') or wall.get('name') or 'unknown'
        quantity = int(safe_float(wall.get('quantity'), 1))
        area = safe_float(wall.get('area_m2'), 0)
        volume = safe_float(wall.get('volume_m3'), 0)

        if wall_type not in statistics_by_type:
            statistics_by_type[wall_type] = {
                'wall_type': wall_type,
                'walls_count': 0,
                'area_m2': 0.0,
                'volume_m3': 0.0,
            }

        statistics_by_type[wall_type]['walls_count'] += quantity
        statistics_by_type[wall_type]['area_m2'] += area * quantity
        statistics_by_type[wall_type]['volume_m3'] += volume * quantity

    wall_types = []
    for wall_type_statistics in statistics_by_type.values():
        wall_type_statistics['area_m2'] = round(wall_type_statistics['area_m2'], 2)
        wall_type_statistics['volume_m3'] = round(wall_type_statistics['volume_m3'], 2)
        wall_types.append(wall_type_statistics)

    return {'wall_types': wall_types}


def form_wall_type_statistics_df(walls_statistics: dict) -> pd.DataFrame:
    rows = []

    for row_num, wall_type_statistics in enumerate(
        walls_statistics.get('wall_types', []),
        start=1,
    ):
        rows.append({
            '№ п/п': row_num,
            'Тип': wall_type_statistics.get('wall_type', ''),
            'Количество': wall_type_statistics.get('walls_count', 0),
            'Площадь, м2': wall_type_statistics.get('area_m2', 0),
            'Объем, м3': wall_type_statistics.get('volume_m3', 0),
        })

    return pd.DataFrame(rows)


def form_result_df(data: dict) -> pd.DataFrame:
    rows = []
    row_num = 1
    
 
    for wall in data.get('walls', []):
        rows.append({
            '№ п/п': row_num, 'Тип (RU)': 'Стены', 'Тип элемента': 'IfcWall',
            'GlobalId': wall.get('id', ''), 'Имя': wall.get('name', ''),
            'Материал': wall.get('material', ''), 'Длина_Width_мм': wall.get('thickness_mm', ''),
            'Глубина_выдавливания_мм': wall.get('thickness_mm', ''),
            'Периметр_мм': wall.get('perimeter_mm', ''), 'Площадь_GrossArea_м2': wall.get('area_m2', ''),
            'Объём_NetVolume_м3': wall.get('volume_m3', ''), 'Количество': wall.get('quantity', 1)
        })
        row_num += 1
    
     
    for col in data.get('columns', []):
        rows.append({
            '№ п/п': row_num, 'Тип (RU)': 'Колонны', 'Тип элемента': 'IfcColumn',
            'GlobalId': col.get('id', ''), 'Имя': col.get('name', ''),
            'Материал': col.get('material', ''), 'Длина_Width_мм': col.get('width_mm', ''),
            'Высота_сечения_мм': col.get('height_mm', ''), 'Объём_NetVolume_м3': col.get('volume_m3', ''),
            'Количество': col.get('quantity', 1)
        })
        row_num += 1
    
     
    for slab in data.get('slabs', []):
        rows.append({
            '№ п/п': row_num, 'Тип (RU)': 'Перекрытия', 'Тип элемента': 'IfcSlab',
            'GlobalId': slab.get('id', ''), 'Имя': slab.get('name', ''),
            'Материал': slab.get('material', ''), 'Глубина_выдавливания_мм': slab.get('thickness_mm', ''),
            'Площадь_GrossArea_м2': slab.get('area_m2', ''), 'Объём_NetVolume_м3': slab.get('volume_m3', ''),
            'Объём_GrossVolume_литры': slab.get('volume_liters', ''), 'Координата_Z_мм': slab.get('elevation_m', ''),
            'Количество': slab.get('quantity', 1)
        })
        row_num += 1
    
 
    for op in data.get('openings', []):
        ifc_type = 'IfcDoor' if op.get('type') == 'дверь' else 'IfcWindow'
        rows.append({
            '№ п/п': row_num, 'Тип (RU)': 'Проёмы', 'Тип элемента': ifc_type,
            'GlobalId': op.get('id', ''), 'Длина_Width_мм': op.get('width_mm', ''),
            'Высота_сечения_мм': op.get('height_mm', ''), 'Количество': op.get('quantity', '')
        })
        row_num += 1
    
 
    for sp in data.get('spaces', []):
        rows.append({
            '№ п/п': row_num, 'Тип (RU)': 'Помещения', 'Тип элемента': 'IfcSpace',
            'GlobalId': sp.get('id', ''), 'Имя': sp.get('name', ''),
            'Площадь_GrossArea_м2': sp.get('area_m2', ''), 'Объём_NetVolume_м3': sp.get('volume_m3', ''),
            'Координата_Z_мм': sp.get('elevation_m', '')
        })
        row_num += 1
    
    return pd.DataFrame(rows)


def save_to_excel(data: dict, filename: str, sheets: dict[str, pd.DataFrame] | None = None) -> str:
    excel_path = os.path.join(OUTPUT_DIR, filename)
    excel_sheets = {"Данные": form_result_df(data)}
    if sheets:
        excel_sheets.update(sheets)

    save_df_to_excel(excel_path, excel_sheets)

    return excel_path

def save_df_to_excel(excel_path, sheets: dict[str, pd.DataFrame]):
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            _write_df_to_excel_sheet(writer, df, sheet_name)

def _write_df_to_excel_sheet(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    for col in worksheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 25)
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font



def safe_float(value, default=0):
     
    if value is None or value == 'не указано' or value == '':
        return default
    try:
        return float(value)
    except:
        return default
    
def save_result(datas):
    if not datas:
        return
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    dfs_for_save = {}
    for i, data in enumerate(datas):
        data = calculate_volumes(data)
        walls_statistics = form_wall_type_statistics(data) # считаем статистику по стенам
        walls_statistics_df = form_wall_type_statistics_df(walls_statistics)

        safe_name = f"drawing_ifc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = os.path.join(OUTPUT_DIR, safe_name)
        dfs_for_save[f"Данные_{i}"] = form_result_df(data)
        dfs_for_save[f"Статистика стен_{i}"] = walls_statistics_df

    save_df_to_excel(excel_path, dfs_for_save)
    return excel_path
