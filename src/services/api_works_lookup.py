"""
Подбор работ через API справочника ТСН (режим КР).

Заменяет фильтрацию работ из data/perechen_kr.xlsx и LLM-отбор
(этапы 1-4) на POST-запросы к эндпоинту
digital-collection/building-elements/positions.

Поток работы:
  1. Выбранные в веб-интерфейсе элементы группируются и преобразуются
     в формат ifc_raw_elements_grouped.json (buildingElementName,
     totalMeasure, characteristics, additionalCharacteristics).
  2. Каждый элемент отправляется отдельным POST-запросом на эндпоинт
     (API принимает один объект, не массив).
  3. Из ответов (data[].works) формируется финальный перечень работ:
     Шифр ТСН = works/code, Наименование расценки/ресурса = works/name,
     Ед. изм. = works/unitOfMeasure. Объём работ рассчитывается по
     totalMeasure соответствующей группы элементов.
  4. Полученные работы отправляются на эндпоинт стоимости
     works/resources (POST), откуда берётся curAll — стоимость одной
     измерительной единицы работы (за 100 м², тонну и т. д. в
     зависимости от okeiValue). Колонка «Стоимость за Ед. Изм.» =
     curAll, «Стоимость» = Объём работ × Стоимость за Ед. Изм.
"""

import base64
import json
import os
import time
from typing import Any, Dict, List, Tuple

import pandas as pd
import requests

from src.core.config import load_config
from src.core.keycloak import KeycloakTokenProvider
from src.core.logger import setup_logger
from src.services.fourth_etap import (
    _get_corrected_volume,
    _add_cost_column,
    add_total_row,
    format_money,
)

logger = setup_logger(__name__)

_cfg = load_config()

API_URL = _cfg.WORKS_API_URL
RESOURCES_API_URL = _cfg.WORKS_RESOURCES_API_URL
API_TOKEN = _cfg.WORKS_API_TOKEN

# Провайдер для автоматического обновления токена (Keycloak, client_credentials).
# Если KEYCLOAK_CLIENT_ID/KEYCLOAK_CLIENT_SECRET не заданы — используется
# статичный WORKS_API_TOKEN как fallback.
if _cfg.KEYCLOAK_CLIENT_ID and _cfg.KEYCLOAK_CLIENT_SECRET:
    _token_provider = KeycloakTokenProvider(
        token_url=_cfg.KEYCLOAK_TOKEN_URL,
        client_id=_cfg.KEYCLOAK_CLIENT_ID,
        client_secret=_cfg.KEYCLOAK_CLIENT_SECRET,
    )
    logger.info(
        "Токен API справочника работ будет получаться автоматически "
        "(Keycloak client_credentials)"
    )
else:
    _token_provider = None
    if not API_TOKEN:
        logger.warning(
            "WORKS_API_TOKEN не задан и Keycloak-клиент не сконфигурирован "
            "(KEYCLOAK_CLIENT_ID/KEYCLOAK_CLIENT_SECRET) — "
            "запросы к API работ будут недоступны"
        )
    else:
        logger.warning(
            "Keycloak-клиент не сконфигурирован (KEYCLOAK_CLIENT_ID/"
            "KEYCLOAK_CLIENT_SECRET) — используется статичный WORKS_API_TOKEN"
        )


def _get_token() -> str:
    """Возвращает актуальный Bearer-токен.

    Если настроен Keycloak-клиент — токен получается и обновляется
    автоматически, иначе возвращается статичный WORKS_API_TOKEN.
    """
    if _token_provider is not None:
        return _token_provider.get_token()
    if not API_TOKEN:
        raise RuntimeError(
            "Не задан Bearer-токен для API подбора работ "
            "(переменная окружения WORKS_API_TOKEN)"
        )
    try:
        payload_b64 = API_TOKEN.split(".")[1]
        payload_b64 += "=" * (-len(payload_b64) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        exp = payload.get("exp")
        exp_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(exp)) if exp else "?"
    except Exception:
        exp_str = "?"
    logger.info(f"Используется статичный WORKS_API_TOKEN (exp={exp_str})")
    return API_TOKEN

# Максимальный размер страницы — чтобы получить все позиции за один запрос.
_PAGE_SIZE = 1000


# =====================================================================
#  POST-ЗАПРОС К API
# =====================================================================

def _fetch_one(element: Dict[str, Any]) -> Dict[str, Any]:
    """POST-запрос для одного элемента.

    API принимает один объект (не массив), поэтому каждый элемент
    отправляется отдельным запросом. Служебные внутренние поля
    (ключи с префиксом '_', например _reinforcementVolumeRatio)
    в запрос не включаются.
    """
    headers = {
        "Authorization": f"Bearer {_get_token()}",
        "Content-Type": "application/json",
    }
    url = f"{API_URL}?pageSize={_PAGE_SIZE}"

    # Полезная нагрузка без внутренних служебных полей группы
    payload = {
        key: value for key, value in element.items()
        if not key.startswith('_')
    }

    def _post() -> requests.Response:
        try:
            return requests.post(url, json=payload, headers=headers, timeout=300)
        except requests.RequestException as exc:
            raise RuntimeError(f"Ошибка вызова API подбора работ: {exc}") from exc

    response = _post()

    # Токен мог быть отозван Keycloak раньше, чем наступил рассчитанный
    # expires_at. В таком случае сбрасываем кеш, получаем новый токен
    # и повторяем запрос ровно один раз.
    if response.status_code == 401 and _token_provider is not None:
        _token_provider.invalidate()
        headers["Authorization"] = f"Bearer {_get_token()}"
        response = _post()

    if response.status_code == 401:
        raise RuntimeError(
            "API подбора работ вернул 401 — Bearer-токен недействителен "
            "(проверьте WORKS_API_TOKEN / KEYCLOAK_CLIENT_ID и "
            "KEYCLOAK_CLIENT_SECRET)"
        )
    if response.status_code >= 400:
        raise RuntimeError(
            f"API подбора работ вернул статус {response.status_code}: "
            f"{response.text[:500]}"
        )

    try:
        data = response.json()
    except ValueError as exc:
        raise RuntimeError(
            f"API подбора работ вернул не-JSON ответ: {response.text[:500]}"
        ) from exc

    if not isinstance(data, dict) or "data" not in data:
        raise RuntimeError(
            f"Некорректный ответ API подбора работ: {str(data)[:500]}"
        )

    return data


def _iter_works(response: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Извлекает список работ из ответа API подбора работ.

    Поддерживает оба формата ответа эндпоинта:
      * вложенный (актуальный): {"data": [{..., "works": [...]}]} —
        позиции справочника с вложенным списком работ;
      * плоский (на случай отката формата): {"data": [work, ...]} —
        список работ на верхнем уровне.

    Аргументы:
        response — разобранный JSON-ответ API.

    Возвращает:
        Плоский список работ (пустой, если работ нет).
    """
    data = response.get("data", []) or []
    works: List[Dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        nested = item.get("works")
        if isinstance(nested, list):
            works.extend(w for w in nested if isinstance(w, dict))
        elif item.get("code") is not None or item.get("name") is not None:
            # Плоский формат: элемент data сам является работой
            works.append(item)
    return works


def fetch_works_from_api(
    elements_json: List[Dict[str, Any]],
) -> List[Tuple[Dict[str, Any], Dict[str, Any]]]:
    """Отправляет элементы в API и собирает ответы.

    API принимает один объект за запрос, поэтому каждый элемент
    из elements_json отправляется отдельным POST-запросом.

    Аргументы:
        elements_json — массив групп элементов в формате
            ifc_raw_elements_grouped.json.

    Возвращает:
        Список пар (element, api_response), где element — исходная
        группа запроса, api_response — ответ API для неё
        ({"data": [...], "paging": {...}}).
    """
    if not elements_json:
        raise RuntimeError("Нет данных для отправки в API подбора работ")

    # Убеждаемся, что токен доступен (из Keycloak-провайдера или WORKS_API_TOKEN).
    _get_token()

    results: List[Tuple[Dict[str, Any], Dict[str, Any]]] = []
    total = len(elements_json)

    for index, element in enumerate(elements_json, 1):
        name = element.get("buildingElementName", "?")
        logger.info(f"API запрос {index}/{total}: {name}")
        response = _fetch_one(element)
        positions = response.get("data", []) or []
        works_count = len(_iter_works(response))
        logger.info(
            f"API ответ {index}/{total}: {name} — "
            f"позиций={len(positions)}, работ={works_count}"
        )
        results.append((element, response))

    return results


# =====================================================================
#  ЗАПРОС СТОИМОСТИ РАБОТ (works/resources)
# =====================================================================

# Максимальное число работ в одном запросе стоимости.
_COST_BATCH_SIZE = 100


def _post_resources(payload: Dict[str, Any]) -> Dict[str, Any]:
    """POST-запрос к эндпоинту стоимости работ (works/resources).

    При 401 сбрасывает кеш токена Keycloak и повторяет запрос один раз.
    """
    headers = {
        "Authorization": f"Bearer {_get_token()}",
        "Content-Type": "application/json",
    }

    def _post() -> requests.Response:
        try:
            return requests.post(
                RESOURCES_API_URL, json=payload, headers=headers, timeout=300
            )
        except requests.RequestException as exc:
            raise RuntimeError(f"Ошибка вызова API стоимости работ: {exc}") from exc

    response = _post()

    if response.status_code == 401 and _token_provider is not None:
        _token_provider.invalidate()
        headers["Authorization"] = f"Bearer {_get_token()}"
        response = _post()

    if response.status_code == 401:
        raise RuntimeError(
            "API стоимости работ вернул 401 — Bearer-токен недействителен "
            "(проверьте WORKS_API_TOKEN / KEYCLOAK_CLIENT_ID и "
            "KEYCLOAK_CLIENT_SECRET)"
        )
    if response.status_code >= 400:
        raise RuntimeError(
            f"API стоимости работ вернул статус {response.status_code}: "
            f"{response.text[:500]}"
        )

    try:
        return response.json()
    except ValueError as exc:
        raise RuntimeError(
            f"API стоимости работ вернул не-JSON ответ: {response.text[:500]}"
        ) from exc


def fetch_work_costs(
    works: List[Dict[str, Any]],
) -> Tuple[Dict[int, float], List[Dict[str, Any]]]:
    """Запрашивает стоимость единицы измерения для списка работ.

    Работы отправляются на эндпоинт works/resources в формате
    {"works": [...], "wrate": false, "coefficient": true}. В ответе для
    каждой работы возвращается curAll — стоимость одной измерительной
    единицы (за 100 м², тонну и т. д. в зависимости от okeiValue).

    Аргументы:
        works — список работ в формате ответа API подбора работ
            (id, code, name, unitOfMeasure, okeiValue, characteristics,
            additionalWorks).

    Возвращает:
        Кортеж (costs, raw_responses), где costs — словарь
        {workId: curAll}, raw_responses — сырые ответы API (для дампа).
    """
    costs: Dict[int, float] = {}
    raw_responses: List[Dict[str, Any]] = []

    # Дедупликация работ по id (одна и та же расценка может прийти
    # в ответах для разных групп элементов)
    unique: Dict[int, Dict[str, Any]] = {}
    for work in works:
        work_id = work.get("id")
        if work_id is None or work_id in unique:
            continue
        unique[work_id] = work

    work_ids = list(unique.keys())
    total = len(work_ids)
    logger.info(f"Запрос стоимости для {total} работ ({RESOURCES_API_URL})")

    for start in range(0, total, _COST_BATCH_SIZE):
        batch_ids = work_ids[start:start + _COST_BATCH_SIZE]
        payload = {
            "works": [unique[wid] for wid in batch_ids],
            "wrate": False,
            "coefficient": True,
        }
        response = _post_resources(payload)
        raw_responses.append(response)

        for cost_work in response.get("works", []) or []:
            work_id = cost_work.get("workId")
            cur_all = cost_work.get("curAll")
            if work_id is None or cur_all is None:
                continue
            costs[work_id] = safe_float(cur_all)

    logger.info(
        f"Получена стоимость для {len(costs)} из {total} работ"
    )
    return costs, raw_responses


# =====================================================================
#  РАСЧЁТ ОБЪЁМА РАБОТ
# =====================================================================

def safe_float(value, default: float = 0.0) -> float:
    """Безопасное преобразование значения во float."""
    if value is None or isinstance(value, bool):
        return default
    try:
        num = float(value)
        if num != num or num in (float("inf"), float("-inf")):  # NaN/Inf
            return default
        return num
    except (ValueError, TypeError):
        return default


def _divisor(okei_value) -> int:
    try:
        okei = int(okei_value)
    except (ValueError, TypeError):
        okei = 1
    return okei if okei and okei > 1 else 1


# Приоритет ключей при выборе «рабочей» площади группы из totalAreas.
# «Площадь, м2» — основная площадь элементов (для стен — боковая/опалубочная,
# для плит/балок — площадь опирания), уже выбранная на этапе подготовки таблицы.
_AREA_KEY_PRIORITY = ['Площадь, м2', 'GrossSideArea', 'NetSideArea', 'GrossArea', 'GROSS']


def _pick_area_value(total_areas: Dict[str, Any]) -> float:
    """Выбирает «рабочую» площадь группы элементов из totalAreas.

    Приоритет ключей: 'Площадь, м2' (основная площадь из сметной таблицы),
    затем боковые/полные площади (GrossSideArea, NetSideArea, GrossArea,
    GROSS). Площадь следа/опирания (Footprint) используется только в крайнем
    случае, если других площадей в группе нет.

    Аргументы:
        total_areas — словарь {имя колонки площади: суммарное значение}.

    Возвращает:
        Число площади (м²) или 0.0, если подходящей площади нет.
    """
    if not total_areas:
        return 0.0
    for priority in _AREA_KEY_PRIORITY:
        for key, value in total_areas.items():
            if priority.lower() in str(key).lower() and 'footprint' not in str(key).lower():
                area_value = safe_float(value)
                if area_value > 0:
                    return area_value
    # Fallback: любая положительная площадь из группы
    for value in total_areas.values():
        area_value = safe_float(value)
        if area_value > 0:
            return area_value
    return 0.0


def _calculate_work_volume(
    work: Dict[str, Any],
    total_measure: Dict[str, Any],
    total_areas: Dict[str, Any] | None = None,
) -> str:
    """Рассчитывает объём работ для расценки из API.

    Аргументы:
        work — расценка из ответа API (unitOfMeasure, okeiValue).
        total_measure — totalMeasure группы элементов запроса
            ({type: volume|area|count, value, unit}).
        total_areas — суммарные площади группы (totalAreas из
            ifc_reference_builder), используются если расценка в м²,
            а основной измеритель группы — объём (например, опалубка
            стен считается по боковой площади при totalMeasure типа volume).

    Возвращает строку с объёмом (или пустую строку, если посчитать нельзя).
    """
    measure_type = (total_measure or {}).get("type", "")
    measure_value = safe_float((total_measure or {}).get("value", 0))

    if measure_value <= 0:
        return ""

    unit = str(work.get("unitOfMeasure", "") or "").lower().replace(" ", "")
    okei_value = work.get("okeiValue")

    is_volume = ("м3" in unit or "m3" in unit or "м[3" in unit or "m[3" in unit)
    is_area = ("м2" in unit or "m2" in unit or "м[2" in unit or "m[2" in unit)
    is_count = "шт" in unit

    if is_volume and measure_type == "volume":
        vol = measure_value / _divisor(okei_value)
        decimals = 4 if _divisor(okei_value) > 1 else 3
        return f"{vol:.{decimals}f}"
    if is_area and measure_type == "area":
        vol = measure_value / _divisor(okei_value)
        decimals = 4 if _divisor(okei_value) > 1 else 2
        return f"{vol:.{decimals}f}"
    if is_count and measure_type == "count":
        vol = measure_value / _divisor(okei_value)
        decimals = 4 if _divisor(okei_value) > 1 else 0
        return f"{vol:.{decimals}f}"

    # Расценка в площади (м²), а основной измеритель группы — объём или прочее.
    # Для таких работ берём суммарную площадь группы из totalAreas
    # (например, монтаж/демонтаж опалубки стен считается по боковой площади).
    if is_area and total_areas:
        area_value = _pick_area_value(total_areas)
        if area_value > 0:
            vol = area_value / _divisor(okei_value)
            decimals = 4 if _divisor(okei_value) > 1 else 2
            return f"{vol:.{decimals}f}"

    # Массовые (т) и прочие измерители посчитать из totalMeasure нельзя.
    return ""

    unit = str(work.get("unitOfMeasure", "") or "").lower().replace(" ", "")
    okei_value = work.get("okeiValue")

    is_volume = ("м3" in unit or "m3" in unit or "м[3" in unit or "m[3" in unit)
    is_area = ("м2" in unit or "m2" in unit or "м[2" in unit or "m[2" in unit)
    is_count = "шт" in unit

    if is_volume and measure_type == "volume":
        vol = measure_value / _divisor(okei_value)
        decimals = 4 if _divisor(okei_value) > 1 else 3
        return f"{vol:.{decimals}f}"
    if is_area and measure_type == "area":
        vol = measure_value / _divisor(okei_value)
        decimals = 4 if _divisor(okei_value) > 1 else 2
        return f"{vol:.{decimals}f}"
    if is_count and measure_type == "count":
        vol = measure_value / _divisor(okei_value)
        decimals = 4 if _divisor(okei_value) > 1 else 0
        return f"{vol:.{decimals}f}"

    # Массовые (т) и прочие измерители посчитать из totalMeasure нельзя.
    return ""


def _resolve_unit_label(work: Dict[str, Any]) -> str:
    """Определяет единицу измерения для колонки «Ед. изм.» финального перечня.

    Обозначение единицы должно соответствовать делителю объёма (okeiValue):
      * без делителя (okeiValue <= 1) — базовая единица: «м³», «м²», «шт»;
      * с делителем (например, 100) — «100 м³», «100 м²», «100 шт»,
        т.е. объём в перечне выражен в сотнях соответствующих единиц.

    Для прочих измерителей (т, кг, м и т.п.) возвращается исходное
    обозначение из API.
    """
    unit = str(work.get("unitOfMeasure", "") or "").lower().replace(" ", "")
    divisor = _divisor(work.get("okeiValue"))

    if "м3" in unit or "m3" in unit or "м[3" in unit or "m[3" in unit:
        base = "м³"
    elif "м2" in unit or "m2" in unit or "м[2" in unit or "m[2" in unit:
        base = "м²"
    elif "шт" in unit:
        base = "шт"
    else:
        return work.get("unitOfMeasure", "") or ""

    return f"{divisor} {base}" if divisor > 1 else base


# =====================================================================
#  ФОРМИРОВАНИЕ ФИНАЛЬНОГО ПЕРЕЧНЯ РАБОТ
# =====================================================================

def build_final_works_from_api(
    api_results: List[Tuple[Dict[str, Any], Dict[str, Any]]],
    output_folder: str,
    save_raw: bool = True,
) -> str:
    """Формирует финальный Excel-перечень работ из ответов API.

    Колонки:
        Шифр ТСН              = works/code
        Наименование расценки/ресурса = works/name
        Ед. изм.              = works/unitOfMeasure
        Объём работ           = по totalMeasure группы элементов запроса
        Стоимость за Ед. Изм. = curAll из API стоимости (works/resources);
                                при недоступности API — цена из price_cost.xlsx
        Стоимость             = Объём работ × Стоимость за Ед. Изм.

    Аргументы:
        api_results   — список пар (element, api_response) из fetch_works_from_api.
        output_folder — папка для сохранения результата.
        save_raw      — сохранять ли сырые ответы API (api_works_response.json).

    Возвращает:
        Путь к созданному файлу ОБЩИЙ_Финальный_перечень_работ.xlsx.
    """
    # Сначала собираем только строки с работами для обработки
    work_rows: List[Dict[str, Any]] = []
    # Все работы из ответов API — для запроса стоимости (works/resources)
    all_works: List[Dict[str, Any]] = []
    # Сохраняем информацию о заголовках элементов и их позициях
    element_info = []  # список словарей с информацией о каждом элементе

    for element_idx, (element, response) in enumerate(api_results):
        # Получаем название элемента из разных возможных полей
        element_name = element.get("buildingElementName") or element.get("name") or f"Элемент {element_idx + 1}"
        
        # Дополнительная информация об элементе
        element_info_parts = []
        
        # Добавляем материал, если есть
        if element.get("characteristics"):
            for char in element["characteristics"]:
                if char.get("name") == "Материал" and char.get("values"):
                    material = char["values"][0].get("strValue", "")
                    if material:
                        element_info_parts.append(material)
        
        # Добавляем имя элемента из Revit (обрезаем ID после двоеточия)
        if element.get("additionalCharacteristics"):
            for char in element["additionalCharacteristics"]:
                if char.get("name") == "Имя элемента" and char.get("values"):
                    revit_name = char["values"][0].get("strValue", "")
                    # Убираем цифровой ID в конце (после последнего двоеточия)
                    if ":" in revit_name:
                        parts = revit_name.split(":")
                        # Если последний элемент - цифры, убираем его
                        if parts[-1].strip().isdigit():
                            revit_name = ":".join(parts[:-1])
                    if revit_name:
                        element_info_parts.append(revit_name)
        
        # Формируем полное название с дополнительной информацией
        element_full_name = element_name
        if element_info_parts:
            element_full_name += f" ({', '.join(element_info_parts)})"
        
        # Добавляем информацию о количестве
        element_count = element.get("elementCount", 1)
        if element_count > 1:
            element_full_name += f" Кол-во: {element_count}"

        total_measure = element.get("totalMeasure", {})
        total_areas = element.get("totalAreas", {}) or {}
        # Суммарный расход арматуры (ReinforcementVolumeRatio, кг) по всем
        # элементам группы. Для работ по установке арматуры пересчитывается
        # в тонны (÷ 1000) и подставляется в «Объём работ».
        reinforcement_ratio = element.get("_reinforcementVolumeRatio")
        positions = response.get("data", []) or []

        # Сохраняем индекс начала работ этого элемента
        start_index = len(work_rows)
        
        for position in positions:
            pos_label = position.get("fullName") or position.get("name", "")

            for work in _iter_works({"data": [position]}):
                volume = _calculate_work_volume(work, total_measure, total_areas)
                work_name = str(work.get("name", ""))
                # Работы по установке арматурных изделий/каркасов/сеток/стержней
                is_rebar_work = "арматур" in work_name.lower()
                if is_rebar_work and reinforcement_ratio:
                    # ReinforcementVolumeRatio в IFC задан в килограммах
                    # (на кубический метр), единица расценки — тонны: кг → т
                    volume = f"{reinforcement_ratio / 1000:.4f}"
                all_works.append(work)
                work_rows.append({
                    "Шифр ТСН": work.get("code", ""),
                    "Наименование расценки/ресурса": work.get("name", ""),
                    "Ед. изм.": _resolve_unit_label(work),
                    "Объём работ": volume,
                    # id работы для сопоставления со стоимостью из API
                    "_workId": work.get("id"),
                })
        
        # Сохраняем информацию об элементе
        element_info.append({
            "header": element_full_name,
            "start_index": start_index,
            "end_index": len(work_rows),
            "is_last": element_idx == len(api_results) - 1
        })

    if not work_rows:
        logger.warning("API подбора работ не вернул ни одной расценки")
        work_rows.append({
            "Шифр ТСН": "",
            "Наименование расценки/ресурса": "Работы не подобраны",
            "Ед. изм.": "",
            "Объём работ": "",
        })

    # Создаем DataFrame только с работами
    df_works = pd.DataFrame(work_rows, columns=[
        "Шифр ТСН", "Наименование расценки/ресурса", "Ед. изм.", "Объём работ",
        "_workId",
    ])

    # Корректировка объёмов по нормам расхода (koefs.xlsx)
    try:
        df_works = _get_corrected_volume(df_works)
    except Exception as exc:
        logger.error(f"Ошибка при корректировке объёма работ: {exc}")

    # Стоимость за единицу измерения — из API цифрового сборника
    # (works/resources, поле curAll). При сбое — fallback на price_cost.xlsx.
    costs: Dict[int, float] = {}
    cost_raw: List[Dict[str, Any]] = []
    try:
        works_for_cost = [w for w in all_works if w.get("id") is not None]
        if works_for_cost:
            costs, cost_raw = fetch_work_costs(works_for_cost)
    except Exception as exc:
        logger.error(f"Ошибка при получении стоимости из API: {exc}")

    if costs:
        df_works["Стоимость за Ед. Изм."] = df_works["_workId"].map(
            lambda wid: costs.get(wid) if wid is not None else None
        )
        df_works["Стоимость за Ед. Изм."] = df_works["Стоимость за Ед. Изм."].apply(
            lambda v: round(v, 2) if safe_float(v) > 0 else ""
        )

        # Стоимость = Объём работ × Стоимость за Ед. Изм.
        def _calc_total(row) -> Any:
            unit_cost = safe_float(row.get("Стоимость за Ед. Изм."), default=0.0)
            volume = safe_float(row.get("Объём работ"), default=0.0)
            if unit_cost > 0 and volume > 0:
                return round(unit_cost * volume, 2)
            return ""

        df_works["Стоимость"] = df_works.apply(_calc_total, axis=1)
    else:
        # Fallback: стоимость из price_cost.xlsx (прежнее поведение)
        logger.warning(
            "Стоимость из API недоступна — используется price_cost.xlsx"
        )
        try:
            df_works = _add_cost_column(df_works)
        except Exception as exc:
            logger.error(f"Ошибка при расчёте стоимости: {exc}")
            df_works["Стоимость"] = ""

    # Форматирование денежных колонок: разряды через пробел,
    # 2 знака после точки (например, '392 458.21')
    if "Стоимость за Ед. Изм." in df_works.columns:
        df_works["Стоимость за Ед. Изм."] = df_works["Стоимость за Ед. Изм."].apply(
            format_money
        )
    if "Стоимость" in df_works.columns:
        df_works["Стоимость"] = df_works["Стоимость"].apply(format_money)

    # Теперь добавляем заголовки элементов в обработанный DataFrame
    final_rows = []
    current_position = 0
    
    for elem_info in element_info:
        # Добавляем заголовок элемента
        final_rows.append({
            "Шифр ТСН": "",
            "Наименование расценки/ресурса": f"{elem_info['header']}",
            "Ед. изм.": "",
            "Объём работ": "",
            "Стоимость за Ед. Изм.": "",
            "Стоимость": "",
            "_is_header": True
        })
        
        # Добавляем работы этого элемента
        for idx in range(elem_info['start_index'], elem_info['end_index']):
            row = df_works.iloc[idx].to_dict()
            row["_is_header"] = False
            final_rows.append(row)
        
        # Добавляем пустую строку после каждого элемента (кроме последнего)
        if not elem_info['is_last']:
            final_rows.append({
                "Шифр ТСН": "",
                "Наименование расценки/ресурса": "",
                "Ед. изм.": "",
                "Объём работ": "",
                "Стоимость за Ед. Изм.": "",
                "Стоимость": "",
                "_is_header": False
            })

    # Создаем финальный DataFrame
    df = pd.DataFrame(final_rows)
    
    # Убедимся, что все нужные колонки есть
    required_columns = [
        "Шифр ТСН", "Наименование расценки/ресурса", "Ед. изм.", "Объём работ",
        "Стоимость за Ед. Изм.", "Стоимость", "_is_header",
    ]
    for col in required_columns:
        if col not in df.columns:
            df[col] = ""
    
    df = df[required_columns]

    # Убираем временную колонку для Excel
    df_for_excel = df.drop(columns=["_is_header"])

    # Итоговая строка: сумма всех значений колонки 'Стоимость'
    df_for_excel = add_total_row(df_for_excel)

    output_path = os.path.join(output_folder, "ОБЩИЙ_Финальный_перечень_работ.xlsx")
    
    # Создаем Excel с форматированием
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_for_excel.to_excel(writer, sheet_name="Данные", index=False)
        
        # Получаем доступ к листу для форматирования
        worksheet = writer.sheets["Данные"]
        
        # Форматируем заголовки элементов (жирный шрифт, заливка)
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Форматируем строки-заголовки
        for row_idx in range(2, len(df) + 2):  # +2 из-за заголовка таблицы
            if df.iloc[row_idx - 2]["_is_header"]:
                for col_idx in range(1, len(df_for_excel.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Настраиваем ширину колонок
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 60
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 18
        worksheet.column_dimensions['F'].width = 15
        
        # Добавляем автофильтр для удобства
        worksheet.auto_filter.ref = f"A1:{chr(64 + len(df_for_excel.columns))}{len(df_for_excel) + 1}"

    logger.info(
        f"Финальный перечень работ из API сохранён: {output_path} "
        f"({len(df_for_excel)} строк)"
    )

    if save_raw:
        raw_path = os.path.join(output_folder, "api_works_response.json")
        combined = {
            "results": [
                {
                    "element": element,
                    "response": response,
                }
                for element, response in api_results
            ],
        }
        with open(raw_path, "w", encoding="utf-8") as fh:
            json.dump(combined, fh, ensure_ascii=False, indent=2, default=str)
        logger.info(f"Сырые ответы API сохранены: {raw_path}")

        if cost_raw:
            cost_path = os.path.join(output_folder, "api_works_cost_response.json")
            with open(cost_path, "w", encoding="utf-8") as fh:
                json.dump(cost_raw, fh, ensure_ascii=False, indent=2, default=str)
            logger.info(f"Сырые ответы API стоимости сохранены: {cost_path}")

    return output_path

